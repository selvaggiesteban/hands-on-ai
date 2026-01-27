#!/usr/bin/env python3
"""
SEO Content Generator - Advanced SEO Analysis and Content Creation Tool
Analyzes websites and generates complete SEO content strategies including:
- Website analysis from sitemaps and URL lists
- Business brief processing and completion
- Competitive research via Google/Bing scraping
- SEO-optimized content generation
- Comprehensive Excel reports with 14 specialized sheets

Supports two operation modes:
1. New Site Mode: Process business brief and generate complete SEO strategy
2. Existing Site Mode: Analyze current site and generate optimization recommendations

Required packages: requests, beautifulsoup4, lxml, openpyxl, python-docx, PyPDF2,
selenium, webdriver-manager, nltk, textblob
"""

import sys
import subprocess

def install_requirements():
    """Install required packages if not available"""
    required_packages = [
        ("requests", "requests>=2.25.1"),
        ("bs4", "beautifulsoup4>=4.9.3"),
        ("lxml", "lxml>=4.6.3"),
        ("openpyxl", "openpyxl>=3.0.0"),
        ("docx", "python-docx>=0.8.11"),
        ("PyPDF2", "PyPDF2>=3.0.1"),
        ("selenium", "selenium>=4.15.0"),
        ("webdriver_manager", "webdriver-manager>=4.0.1"),
        ("nltk", "nltk>=3.8.1"),
        ("textblob", "textblob>=0.17.1")
    ]

    missing_packages = []

    for import_name, package_name in required_packages:
        try:
            __import__(import_name)
        except ImportError:
            missing_packages.append(package_name)

    if missing_packages:
        print(f"Installing missing packages: {', '.join(missing_packages)}")
        try:
            subprocess.check_call([
                sys.executable, "-m", "pip", "install"
            ] + missing_packages)
            print("Packages installed successfully!")
        except subprocess.CalledProcessError as e:
            print(f"Error installing packages: {e}")
            print("Please install manually with: pip install " + " ".join(missing_packages))
            sys.exit(1)

# Install requirements before importing other modules
install_requirements()

import requests
import urllib3
import xml.etree.ElementTree as ET
import argparse
import os
import time
from urllib.parse import urlparse, urljoin

# Disable SSL warnings for sites with certificate errors
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from bs4 import BeautifulSoup
from collections import Counter, defaultdict
from datetime import datetime
import re
import json
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell

# New imports for enhanced functionality
from docx import Document
import PyPDF2
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import nltk
from textblob import TextBlob
import random
from pathlib import Path

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class SEOContentGenerator:
    def __init__(self, delay=1):
        self.delay = delay
        self.analyzed_urls = []
        self.content_data = []
        self.keywords = Counter()
        self.topics = defaultdict(list)
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        # Disable SSL certificate verification for sites with certificate errors
        self.session.verify = False

        # Cache configuration
        self.cache_dir = Path('.seo_cache')
        self.cache_dir.mkdir(exist_ok=True)
        self.cache_rescrape_minutes = 60   # Re-scrape every: 60 minutes (OBLIGATORIO)
        self.cache_cleanup_hours = 24      # Delete cache after: 24 hours
        self.current_domain = None         # Current domain being analyzed
        self.domain_cache_dir = None       # Cache directory for current domain

        # New properties for content generation
        self.operation_mode = None  # 'new_site' or 'existing_site'
        self.business_data = {}
        self.brief_template = self._load_brief_template()
        self.competitive_analysis = {}
        self.generated_content = {}
        self.serp_data = {}
        self.driver = None  # Selenium WebDriver

        # URL Classification and Filtering
        self.url_classifications = {}  # Store URL type classifications
        self.site_ctas = []  # Store detected CTAs from site

    def _load_brief_template(self):
        """Load the embedded brief template"""
        return {
            'general_info': {
                'nombre_comercial': '',
                'descripcion_corta': '',
                'actividad_principal': '',
                'cobertura_geografica': '',
                'medios_pago': '',
                'diferenciales': [],
                'objetivo_principal': '',
                'cta_global': ''
            },
            'locations': [],
            'social_media': {
                'instagram': '',
                'facebook': '',
                'linkedin': '',
                'twitter': '',
                'pinterest': '',
                'tiktok': ''
            },
            'brand': {
                'logo': '',
                'colores': '',
                'tipografias': ''
            },
            'pages': {
                'home': {},
                'sobre_nosotros': {},
                'servicios': {},
                'blog': {},
                'contacto': {}
            },
            'blog_posts': []
        }

    def _should_exclude_url(self, url):
        """Filter out WordPress/Elementor non-content URLs"""
        url_lower = url.lower()

        # Exclude patterns (WordPress, Elementor, taxonomies, archives)
        exclude_patterns = [
            '?elementor_library=',
            '/elementor-',
            '/category/',
            '/tag/',
            '/author/',
            '/search',
            '/page/',
            '/feed/',
            '/wp-json/',
            '/wp-content/',
            '/wp-includes/',
            'sitemap',
            '/attachment/',
            '?p=',
            '?page_id=',
            '/trackback',
            '/comment-page-',
            '/print/',
            '/embed/',
            '?replytocom='
        ]

        # Check if URL matches any exclude pattern
        for pattern in exclude_patterns:
            if pattern in url_lower:
                return True

        # Exclude archive patterns like /2025/04/ without meaningful slug
        archive_pattern = r'/\d{4}/\d{2}/?$'
        if re.search(archive_pattern, url):
            return True

        return False

    def _classify_url_type(self, url):
        """Classify URL by type based on friendly URL patterns"""
        url_lower = url.lower()
        parsed = urlparse(url)
        path = parsed.path.lower()

        # Homepage
        if path in ['/', '', '/index.html', '/index.php']:
            return 'homepage'

        # Blog/Articles - improved detection
        blog_indicators = ['/blog/', '/articulo/', '/post/', '/noticia/', '/news/']
        if any(indicator in path for indicator in blog_indicators):
            return 'blog'

        # Date pattern in URL (typically blog posts) like /2024/05/post-title
        if re.search(r'/\d{4}/\d{2}/.+', path):
            return 'blog'

        # Services
        service_indicators = ['/servicio', '/service', '/tratamiento', '/treatment', '/especialidad']
        if any(indicator in path for indicator in service_indicators):
            return 'service'

        # Products
        product_indicators = ['/producto', '/product', '/tienda', '/shop', '/catalog']
        if any(indicator in path for indicator in product_indicators):
            return 'product'

        # About/Company pages
        about_indicators = ['/sobre-', '/nosotros', '/about', '/empresa', '/quienes', '/equipo', '/team']
        if any(indicator in path for indicator in about_indicators):
            return 'about'

        # Contact
        contact_indicators = ['/contacto', '/contact', '/cita', '/appointment', '/reserva']
        if any(indicator in path for indicator in contact_indicators):
            return 'contact'

        # Legal pages
        legal_indicators = ['/aviso', '/legal', '/privacy', '/privacidad', '/terminos', '/terms', '/cookies']
        if any(indicator in path for indicator in legal_indicators):
            return 'legal'

        # FAQ/Help
        if any(word in path for word in ['/faq', '/ayuda', '/help', '/preguntas']):
            return 'faq'

        # Default: general page
        return 'page'

    def _classify_all_urls(self, urls):
        """Classify all URLs and store in url_classifications"""
        classified = {
            'homepage': [],
            'blog': [],
            'service': [],
            'product': [],
            'about': [],
            'contact': [],
            'faq': [],
            'legal': [],
            'page': []
        }

        for url in urls:
            if not self._should_exclude_url(url):
                url_type = self._classify_url_type(url)
                classified[url_type].append(url)
                self.url_classifications[url] = url_type

        # Log classification summary
        logger.info("📊 URL Classification Summary:")
        for url_type, url_list in classified.items():
            if url_list:
                logger.info(f"  - {url_type.title()}: {len(url_list)} URLs")

        return classified

    def _get_industry_knowledge_banks(self):
        """Knowledge banks for 20+ industries with common questions, services, and terminology"""
        return {
            'fisioterapia': {
                'keywords': ['fisioterapia', 'fisioterapeuta', 'rehabilitación', 'lesión', 'dolor', 'tratamiento', 'terapia', 'recuperación'],
                'services': ['rehabilitación', 'masajes', 'electroterapia', 'ejercicios terapéuticos', 'vendaje funcional', 'punción seca'],
                'common_questions': [
                    '¿Cuántas sesiones necesito?',
                    '¿Cuánto dura una sesión de fisioterapia?',
                    '¿El tratamiento es doloroso?',
                    '¿Cuánto tiempo tarda la recuperación?',
                    '¿Aceptan seguro médico?'
                ],
                'h2_templates': [
                    'Tipos de {keyword} que tratamos',
                    'Proceso de recuperación y rehabilitación',
                    'Técnicas de {keyword} avanzada',
                    'Beneficios del tratamiento {keyword}',
                    'Precio y coste de las sesiones'
                ]
            },
            'dentista': {
                'keywords': ['dental', 'dentista', 'odontología', 'dientes', 'sonrisa', 'ortodoncia', 'implante'],
                'services': ['ortodoncia', 'implantes', 'blanqueamiento', 'endodoncia', 'periodoncia', 'cirugía oral'],
                'common_questions': [
                    '¿Duele el tratamiento dental?',
                    '¿Cuánto cuesta un implante dental?',
                    '¿Cada cuánto debo ir al dentista?',
                    '¿Cuánto dura el blanqueamiento dental?',
                    '¿Aceptan financiación?'
                ],
                'h2_templates': [
                    'Tratamientos de {keyword} disponibles',
                    'Precio de {keyword} en [ciudad]',
                    'Tecnología dental de última generación',
                    '¿Por qué elegir nuestra clínica dental?',
                    'Primera consulta gratuita'
                ]
            },
            'abogado': {
                'keywords': ['abogado', 'legal', 'jurídico', 'derecho', 'abogacía', 'asesoría', 'bufete'],
                'services': ['derecho civil', 'derecho penal', 'derecho laboral', 'divorcio', 'herencias', 'reclamaciones'],
                'common_questions': [
                    '¿Cuánto cobra un abogado?',
                    '¿Cuánto dura el proceso legal?',
                    '¿Ofrecen primera consulta gratuita?',
                    '¿Qué documentos necesito?',
                    '¿Trabajan con tarifa plana?'
                ],
                'h2_templates': [
                    'Especialización en {keyword}',
                    'Casos de éxito en {keyword}',
                    'Proceso y plazos del trámite legal',
                    'Honorarios y formas de pago',
                    'Contacta con nuestro despacho'
                ]
            },
            'restaurante': {
                'keywords': ['restaurante', 'comida', 'gastronomía', 'menú', 'cocina', 'chef', 'reserva'],
                'services': ['comida a domicilio', 'eventos privados', 'menú del día', 'catering', 'reservas'],
                'common_questions': [
                    '¿Cuál es el horario del restaurante?',
                    '¿Tienen menú del día?',
                    '¿Se puede reservar mesa?',
                    '¿Tienen opciones veganas/vegetarianas?',
                    '¿Hacen comida para llevar?'
                ],
                'h2_templates': [
                    'Nuestra carta y especialidades',
                    'Menú del día {keyword}',
                    'Reserva tu mesa en [restaurante]',
                    'Eventos y celebraciones privadas',
                    'Delivery y comida a domicilio'
                ]
            },
            'gimnasio': {
                'keywords': ['gimnasio', 'fitness', 'entrenamiento', 'crossfit', 'musculación', 'cardio'],
                'services': ['entrenamiento personal', 'clases dirigidas', 'nutrición', 'spinning', 'yoga', 'pilates'],
                'common_questions': [
                    '¿Cuánto cuesta la cuota mensual?',
                    '¿Tienen entrenador personal?',
                    '¿Cuál es el horario del gimnasio?',
                    '¿Hay clases dirigidas incluidas?',
                    '¿Ofrecen clase de prueba gratuita?'
                ],
                'h2_templates': [
                    'Instalaciones y equipamiento del gimnasio',
                    'Clases dirigidas de {keyword}',
                    'Entrenamiento personalizado',
                    'Planes y tarifas mensuales',
                    'Horarios y ubicación'
                ]
            },
            'psicologia': {
                'keywords': ['psicólogo', 'psicología', 'terapia', 'ansiedad', 'depresión', 'coaching'],
                'services': ['terapia individual', 'terapia de pareja', 'terapia infantil', 'coaching', 'ansiedad', 'depresión'],
                'common_questions': [
                    '¿Cuánto dura una sesión de psicología?',
                    '¿Cuántas sesiones voy a necesitar?',
                    '¿Es confidencial la terapia?',
                    '¿Hacen terapia online?',
                    '¿Cuánto cuesta una consulta?'
                ],
                'h2_templates': [
                    'Tipos de terapia {keyword}',
                    'Tratamiento para {keyword}',
                    'Beneficios de la terapia psicológica',
                    'Metodología y enfoque terapéutico',
                    'Reserva tu primera consulta'
                ]
            },
            'inmobiliaria': {
                'keywords': ['inmobiliaria', 'piso', 'casa', 'alquiler', 'venta', 'vivienda', 'propiedad'],
                'services': ['venta de pisos', 'alquiler', 'tasación', 'gestión de inmuebles', 'inversión'],
                'common_questions': [
                    '¿Cuánto cobran de comisión?',
                    '¿Tienen pisos en alquiler?',
                    '¿Hacen tasaciones gratuitas?',
                    '¿Cuánto tiempo tarda en venderse una casa?',
                    '¿Ayudan con la hipoteca?'
                ],
                'h2_templates': [
                    'Pisos en venta en [ciudad]',
                    'Alquiler de viviendas {keyword}',
                    'Proceso de compra paso a paso',
                    'Tasación gratuita de tu propiedad',
                    'Por qué elegirnos como tu inmobiliaria'
                ]
            },
            'veterinario': {
                'keywords': ['veterinario', 'veterinaria', 'mascota', 'perro', 'gato', 'animal'],
                'services': ['consulta veterinaria', 'vacunación', 'cirugía', 'urgencias', 'peluquería canina'],
                'common_questions': [
                    '¿Cuánto cuesta la consulta veterinaria?',
                    '¿Tienen servicio de urgencias?',
                    '¿Hacen cirugías?',
                    '¿Cuándo debo vacunar a mi mascota?',
                    '¿Atienden urgencias 24 horas?'
                ],
                'h2_templates': [
                    'Servicios veterinarios para {keyword}',
                    'Urgencias veterinarias 24h',
                    'Vacunación y prevención',
                    'Precio de consulta y tratamientos',
                    'Cuidado integral de tu mascota'
                ]
            },
            'hotel': {
                'keywords': ['hotel', 'alojamiento', 'habitación', 'hospedaje', 'turismo', 'resort'],
                'services': ['habitaciones', 'spa', 'restaurante', 'piscina', 'eventos', 'wifi'],
                'common_questions': [
                    '¿Cuál es el horario de check-in?',
                    '¿Tienen wifi gratuito?',
                    '¿Admiten mascotas?',
                    '¿El desayuno está incluido?',
                    '¿Se puede cancelar la reserva?'
                ],
                'h2_templates': [
                    'Habitaciones y suites disponibles',
                    'Servicios e instalaciones del hotel',
                    'Ubicación y puntos de interés cercanos',
                    'Tarifas y ofertas especiales',
                    'Reserva tu estancia en [hotel]'
                ]
            },
            'taller_mecanico': {
                'keywords': ['taller', 'mecánico', 'reparación', 'coche', 'vehículo', 'auto'],
                'services': ['revisión', 'reparación motor', 'cambio aceite', 'neumáticos', 'ITV', 'chapa y pintura'],
                'common_questions': [
                    '¿Cuánto cuesta la revisión del coche?',
                    '¿Hacen presupuesto gratuito?',
                    '¿Cuánto tarda la reparación?',
                    '¿Ofrecen servicio de grúa?',
                    '¿Tienen coche de sustitución?'
                ],
                'h2_templates': [
                    'Servicios de reparación {keyword}',
                    'Revisión completa del vehículo',
                    'Presupuesto sin compromiso',
                    'Precio y tiempo de reparación',
                    'Por qué confiar en nuestro taller'
                ]
            },
            'academia': {
                'keywords': ['academia', 'clases', 'formación', 'curso', 'profesor', 'aprender'],
                'services': ['clases particulares', 'preparación exámenes', 'refuerzo escolar', 'idiomas', 'online'],
                'common_questions': [
                    '¿Cuánto cuestan las clases?',
                    '¿Tienen clases online?',
                    '¿Cuál es el horario?',
                    '¿Hacen preparación de exámenes?',
                    '¿Ofrecen clase de prueba gratuita?'
                ],
                'h2_templates': [
                    'Cursos y clases de {keyword}',
                    'Metodología de enseñanza',
                    'Horarios y grupos disponibles',
                    'Precio de matrícula y mensualidad',
                    'Inscríbete en nuestra academia'
                ]
            },
            'estetica': {
                'keywords': ['estética', 'belleza', 'tratamiento', 'facial', 'corporal', 'depilación'],
                'services': ['depilación láser', 'tratamientos faciales', 'masajes', 'manicura', 'peeling'],
                'common_questions': [
                    '¿Cuánto cuesta la depilación láser?',
                    '¿Cuántas sesiones necesito?',
                    '¿Es doloroso el tratamiento?',
                    '¿Tienen promociones?',
                    '¿Se puede financiar?'
                ],
                'h2_templates': [
                    'Tratamientos de {keyword} avanzados',
                    'Depilación láser definitiva',
                    'Precio y paquetes de sesiones',
                    'Tecnología y equipamiento',
                    'Reserva tu cita de valoración'
                ]
            },
            'peluqueria': {
                'keywords': ['peluquería', 'peluquero', 'corte', 'pelo', 'tinte', 'peinado'],
                'services': ['corte', 'color', 'mechas', 'alisado', 'tratamiento capilar', 'peinado'],
                'common_questions': [
                    '¿Cuánto cuesta un corte de pelo?',
                    '¿Necesito cita previa?',
                    '¿Hacen alisado brasileño?',
                    '¿Tienen tratamientos para el cabello?',
                    '¿Cuál es el precio del tinte?'
                ],
                'h2_templates': [
                    'Servicios de peluquería y belleza',
                    'Corte y color de pelo',
                    'Tratamientos capilares profesionales',
                    'Precio de servicios',
                    'Reserva tu cita online'
                ]
            },
            'consultoria': {
                'keywords': ['consultoría', 'consultor', 'asesoría', 'negocio', 'empresa', 'estrategia'],
                'services': ['consultoría estratégica', 'análisis de negocio', 'transformación digital', 'auditoría'],
                'common_questions': [
                    '¿Cómo funciona la consultoría?',
                    '¿Cuánto cobran?',
                    '¿Cuánto dura el proceso?',
                    '¿En qué sectores trabajan?',
                    '¿Ofrecen primera consulta gratuita?'
                ],
                'h2_templates': [
                    'Servicios de consultoría {keyword}',
                    'Metodología y proceso de trabajo',
                    'Casos de éxito y resultados',
                    'Sectores en los que trabajamos',
                    'Contacta con nuestros consultores'
                ]
            },
            'arquitectura': {
                'keywords': ['arquitecto', 'arquitectura', 'proyecto', 'diseño', 'construcción', 'reforma'],
                'services': ['proyectos arquitectónicos', 'reformas', 'interiorismo', 'planos', 'licencias'],
                'common_questions': [
                    '¿Cuánto cobra un arquitecto?',
                    '¿Cuánto tiempo tarda el proyecto?',
                    '¿Gestionan las licencias?',
                    '¿Hacen reformas integrales?',
                    '¿Ofrecen servicio de interiorismo?'
                ],
                'h2_templates': [
                    'Proyectos de {keyword} a medida',
                    'Reformas integrales y rehabilitación',
                    'Proceso de diseño arquitectónico',
                    'Honorarios y presupuesto',
                    'Solicita tu consulta gratuita'
                ]
            },
            'fotografo': {
                'keywords': ['fotógrafo', 'fotografía', 'fotos', 'sesión', 'boda', 'reportaje'],
                'services': ['bodas', 'reportajes', 'sesiones familiares', 'fotografía corporativa', 'eventos'],
                'common_questions': [
                    '¿Cuánto cuesta una sesión de fotos?',
                    '¿Cuándo recibo las fotos?',
                    '¿Incluyen retoque?',
                    '¿Cuántas fotos entregáis?',
                    '¿Hacen fotografía de bodas?'
                ],
                'h2_templates': [
                    'Servicios de {keyword} profesional',
                    'Paquetes y precios de sesiones',
                    'Portfolio de trabajos realizados',
                    '¿Qué incluye la sesión fotográfica?',
                    'Reserva tu sesión de fotos'
                ]
            },
            'contabilidad': {
                'keywords': ['gestoría', 'contabilidad', 'fiscal', 'laboral', 'contable', 'asesoría'],
                'services': ['gestión fiscal', 'gestión laboral', 'contabilidad', 'declaraciones', 'nóminas'],
                'common_questions': [
                    '¿Cuánto cuesta una gestoría?',
                    '¿Qué servicios incluye?',
                    '¿Gestionan la declaración de la renta?',
                    '¿Atienden autónomos?',
                    '¿Cuánto tardan en procesar un trámite?'
                ],
                'h2_templates': [
                    'Servicios de gestoría y {keyword}',
                    'Asesoría fiscal para autónomos y empresas',
                    'Gestión laboral y nóminas',
                    'Tarifas y cuotas mensuales',
                    'Solicita presupuesto sin compromiso'
                ]
            },
            'nutricionista': {
                'keywords': ['nutrición', 'nutricionista', 'dieta', 'alimentación', 'peso', 'adelgazar'],
                'services': ['dieta personalizada', 'nutrición deportiva', 'pérdida de peso', 'educación nutricional'],
                'common_questions': [
                    '¿Cuánto cuesta la consulta de nutrición?',
                    '¿Cuántas visitas necesito?',
                    '¿Hacen dietas personalizadas?',
                    '¿Atienden nutrición deportiva?',
                    '¿Tienen consulta online?'
                ],
                'h2_templates': [
                    'Servicios de {keyword} personalizados',
                    'Plan nutricional a medida',
                    'Pérdida de peso saludable',
                    'Precio y seguimiento nutricional',
                    'Pide tu cita con nutricionista'
                ]
            },
            'fontanero': {
                'keywords': ['fontanero', 'fontanería', 'instalación', 'reparación', 'fuga', 'tubería'],
                'services': ['reparaciones', 'instalación', 'urgencias', 'desatascos', 'calefacción'],
                'common_questions': [
                    '¿Cuánto cobra un fontanero?',
                    '¿Tienen servicio de urgencias?',
                    '¿Cuánto tardan en venir?',
                    '¿Hacen presupuesto gratuito?',
                    '¿Atienden los fines de semana?'
                ],
                'h2_templates': [
                    'Servicios de {keyword} profesional',
                    'Urgencias 24 horas disponibles',
                    'Reparaciones e instalaciones',
                    'Precio y presupuesto sin compromiso',
                    'Llama ahora y resolvemos tu problema'
                ]
            },
            'electricista': {
                'keywords': ['electricista', 'electricidad', 'instalación', 'luz', 'cuadro eléctrico'],
                'services': ['instalaciones eléctricas', 'reparaciones', 'cuadros eléctricos', 'certificados', 'urgencias'],
                'common_questions': [
                    '¿Cuánto cobra un electricista?',
                    '¿Hacen certificados eléctricos?',
                    '¿Atienden urgencias?',
                    '¿Hacen presupuesto gratuito?',
                    '¿Cuánto tiempo tarda la instalación?'
                ],
                'h2_templates': [
                    'Servicios de instalaciones {keyword}',
                    'Reparaciones eléctricas urgentes',
                    'Certificados y boletines eléctricos',
                    'Precio y presupuesto de trabajos',
                    'Contacta con electricistas profesionales'
                ]
            }
        }

    def _detect_business_niche(self, page):
        """
        Detect business niche/industry from URL structure, title, and content

        Args:
            page: Dict with 'url', 'title', 'h1', 'body_text', 'meta_description'

        Returns:
            tuple: (niche_name, confidence_score)
        """
        url = page.get('url', '').lower()
        title = page.get('title', '').lower()
        h1 = page.get('h1', [''])[0].lower() if page.get('h1') else ''
        meta = page.get('meta_description', '').lower()
        body = page.get('body_text', '').lower()

        # Combine all text for analysis
        all_text = f"{url} {title} {h1} {meta} {body}"

        industry_banks = self._get_industry_knowledge_banks()

        # Score each industry
        scores = {}
        for niche, data in industry_banks.items():
            score = 0
            keywords = data['keywords']

            # Check presence of industry keywords
            for keyword in keywords:
                if keyword in url:
                    score += 10  # URL is strongest signal
                if keyword in title:
                    score += 8
                if keyword in h1:
                    score += 6
                if keyword in meta:
                    score += 4
                # Count occurrences in body (cap at 5)
                body_count = min(body.count(keyword), 5)
                score += body_count * 2

            scores[niche] = score

        # Get top niche
        if not scores or max(scores.values()) == 0:
            return ('general', 0)

        top_niche = max(scores, key=scores.get)
        confidence = scores[top_niche]

        return (top_niche, confidence)

    def _parse_url_structure(self, url):
        """
        Parse URL structure to extract semantic meaning from path segments

        Args:
            url: Full URL string

        Returns:
            dict: {
                'section': First path segment (servicios, blog, etc.),
                'category': Second path segment,
                'subcategory': Third path segment,
                'slug': Final segment,
                'segments': All segments as list
            }
        """
        parsed = urlparse(url)
        path = parsed.path.strip('/')

        if not path:
            return {
                'section': 'homepage',
                'category': None,
                'subcategory': None,
                'slug': 'homepage',
                'segments': []
            }

        segments = path.split('/')

        return {
            'section': segments[0] if len(segments) > 0 else None,
            'category': segments[1] if len(segments) > 1 else None,
            'subcategory': segments[2] if len(segments) > 2 else None,
            'slug': segments[-1],  # Last segment
            'segments': segments
        }

    def select_operation_mode(self):
        """Interactive mode selection"""
        print("\n" + "="*60)
        print("🚀 SEO CONTENT GENERATOR")
        print("="*60)
        print("\nSeleccione el modo de operación:")
        print("\n1. 📄 SITIO NUEVO - Trabajar con brief de negocio")
        print("   • Procesar archivo de brief (.txt, .docx, .pdf)")
        print("   • Generar estrategia SEO completa")
        print("   • Crear contenido desde cero")
        print("\n2. 🔍 SITIO EXISTENTE - Analizar sitio actual")
        print("   • Analizar sitemap existente")
        print("   • Optimizar páginas seleccionadas")
        print("   • Mejorar contenido actual")
        print("\n" + "-"*60)

        while True:
            try:
                choice = input("\nIngrese su opción (1 o 2): ").strip()
                if choice == '1':
                    self.operation_mode = 'new_site'
                    print("\n✅ Modo seleccionado: SITIO NUEVO")
                    return 'new_site'
                elif choice == '2':
                    self.operation_mode = 'existing_site'
                    print("\n✅ Modo seleccionado: SITIO EXISTENTE")
                    return 'existing_site'
                else:
                    print("❌ Opción inválida. Por favor ingrese 1 o 2.")
            except KeyboardInterrupt:
                print("\n\n👋 Operación cancelada por el usuario.")
                sys.exit(0)

    def process_brief_file(self, file_path=None):
        """Process business brief from txt, docx, or pdf file"""
        if not file_path:
            print("\n📄 PROCESAMIENTO DE BRIEF DE NEGOCIO")
            print("-" * 50)
            file_path = input("Ingrese la ruta del archivo de brief (.txt, .docx, .pdf): ").strip().strip('"')

        if not os.path.exists(file_path):
            print(f"❌ Error: El archivo '{file_path}' no existe.")
            return None

        file_ext = Path(file_path).suffix.lower()
        content = ""

        try:
            if file_ext == '.txt':
                content = self._read_txt_file(file_path)
            elif file_ext == '.docx':
                content = self._read_docx_file(file_path)
            elif file_ext == '.pdf':
                content = self._read_pdf_file(file_path)
            else:
                print(f"❌ Formato de archivo no soportado: {file_ext}")
                print("📋 Formatos soportados: .txt, .docx, .pdf")
                return None

            logger.info(f"Brief loaded from: {file_path}")
            print(f"✅ Brief cargado exitosamente desde: {os.path.basename(file_path)}")

            # Parse brief content
            self.business_data = self._parse_brief_content(content)
            return content

        except Exception as e:
            logger.error(f"Error processing brief file: {e}")
            print(f"❌ Error procesando el archivo: {e}")
            return None

    def _read_txt_file(self, file_path):
        """Read text file"""
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()

    def _read_docx_file(self, file_path):
        """Read Word document"""
        doc = Document(file_path)
        content = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                content.append(paragraph.text.strip())
        return '\n'.join(content)

    def _read_pdf_file(self, file_path):
        """Read PDF file"""
        content = []
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text = page.extract_text()
                if text.strip():
                    content.append(text.strip())
        return '\n'.join(content)

    def _parse_brief_content(self, content):
        """Parse brief content and extract business data"""
        business_data = self.brief_template.copy()

        # Basic parsing using regex patterns
        patterns = {
            'nombre_comercial': r'nombre\s*comercial[:\-]?\s*(.+?)(?:\n|$)',
            'descripcion_corta': r'descripci[óo]n\s*corta[:\-]?\s*(.+?)(?:\n|$)',
            'actividad_principal': r'actividad\s*principal[:\-]?\s*(.+?)(?:\n|$)',
            'cobertura_geografica': r'cobertura\s*geogr[áa]fica[:\-]?\s*(.+?)(?:\n|$)',
            'objetivo_principal': r'objetivo\s*principal[:\-]?\s*(.+?)(?:\n|$)',
        }

        content_lower = content.lower()

        for field, pattern in patterns.items():
            match = re.search(pattern, content_lower, re.IGNORECASE | re.MULTILINE)
            if match:
                business_data['general_info'][field] = match.group(1).strip()

        # Extract social media links
        social_patterns = {
            'instagram': r'instagram[:\-]?\s*(.+?)(?:\n|$)',
            'facebook': r'facebook[:\-]?\s*(.+?)(?:\n|$)',
            'linkedin': r'linkedin[:\-]?\s*(.+?)(?:\n|$)',
        }

        for platform, pattern in social_patterns.items():
            match = re.search(pattern, content_lower, re.IGNORECASE | re.MULTILINE)
            if match:
                business_data['social_media'][platform] = match.group(1).strip()

        return business_data

    def validate_brief_completeness(self):
        """Validate brief completeness and request missing data"""
        print("\n🔍 VALIDACIÓN DE COMPLETITUD DEL BRIEF")
        print("-" * 50)

        missing_fields = []
        required_fields = {
            'general_info': {
                'nombre_comercial': 'Nombre comercial del negocio',
                'actividad_principal': 'Actividad principal del negocio',
                'cobertura_geografica': 'Cobertura geográfica (ciudades/barrios)',
                'objetivo_principal': 'Objetivo principal del sitio web'
            }
        }

        # Check missing required fields
        for section, fields in required_fields.items():
            for field, description in fields.items():
                if not self.business_data.get(section, {}).get(field):
                    missing_fields.append((section, field, description))

        if missing_fields:
            print(f"📋 Se detectaron {len(missing_fields)} campos obligatorios faltantes:")
            self.request_missing_data(missing_fields)
        else:
            print("✅ El brief está completo con todos los campos obligatorios.")

        # Optional fields with smart defaults
        self._fill_optional_fields()

    def request_missing_data(self, missing_fields):
        """Interactively request missing data from user"""
        print("\n📝 SOLICITUD DE DATOS FALTANTES")
        print("-" * 50)

        for section, field, description in missing_fields:
            while True:
                try:
                    value = input(f"\n• {description}: ").strip()
                    if value:
                        if section not in self.business_data:
                            self.business_data[section] = {}
                        self.business_data[section][field] = value
                        print(f"  ✅ Guardado: {value}")
                        break
                    else:
                        print("  ⚠️  Este campo es obligatorio. Por favor ingrese un valor.")
                except KeyboardInterrupt:
                    print("\n\n👋 Operación cancelada por el usuario.")
                    sys.exit(0)

        print(f"\n✅ Todos los campos obligatorios han sido completados.")

    def _fill_optional_fields(self):
        """Fill optional fields with smart defaults if missing"""
        defaults = {
            'general_info': {
                'descripcion_corta': f"Servicios profesionales de {self.business_data.get('general_info', {}).get('actividad_principal', 'nuestro negocio')}",
                'medios_pago': 'Efectivo, tarjetas de crédito/débito, transferencias bancarias',
                'cta_global': 'Contáctanos'
            }
        }

        for section, fields in defaults.items():
            for field, default_value in fields.items():
                if not self.business_data.get(section, {}).get(field):
                    if section not in self.business_data:
                        self.business_data[section] = {}
                    self.business_data[section][field] = default_value

    def setup_webdriver(self):
        """Setup Selenium WebDriver with Chrome"""
        try:
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--window-size=1920,1080')
            chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
            # Ignore SSL certificate errors
            chrome_options.add_argument('--ignore-certificate-errors')
            chrome_options.add_argument('--allow-insecure-localhost')

            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            logger.info("WebDriver initialized successfully")
            return True

        except Exception as e:
            logger.error(f"Error setting up WebDriver: {e}")
            print(f"⚠️  No se pudo inicializar el navegador. Usando modo básico sin investigación competitiva.")
            return False

    def search_competitors(self, keywords):
        """Perform competitive research via Google and Bing"""
        if not self.driver and not self.setup_webdriver():
            return {}

        print(f"\n🔍 INVESTIGACIÓN COMPETITIVA")
        print("-" * 50)
        print(f"🎯 Analizando keywords: {', '.join(keywords[:3])}...")

        all_competitors = {}
        search_engines = ['google', 'bing']

        for engine in search_engines:
            print(f"\n📊 Buscando en {engine.upper()}...")

            for keyword in keywords[:3]:  # Limit to top 3 keywords
                try:
                    competitors = self._search_single_keyword(engine, keyword)
                    if competitors:
                        all_competitors[f"{engine}_{keyword}"] = competitors
                        print(f"  ✅ {keyword}: {len(competitors)} competidores encontrados")

                    # Delay between searches
                    time.sleep(random.uniform(2, 4))

                except Exception as e:
                    logger.error(f"Error searching {keyword} on {engine}: {e}")
                    print(f"  ❌ Error buscando '{keyword}' en {engine}")

        self.competitive_analysis = all_competitors
        return all_competitors

    def _search_single_keyword(self, engine, keyword):
        """Search single keyword on specified engine"""
        competitors = []
        max_pages = 3  # Reduced from 5 for efficiency

        for page in range(max_pages):
            try:
                if engine == 'google':
                    url = f"https://www.google.com/search?q={keyword}&start={page * 10}"
                else:  # bing
                    url = f"https://www.bing.com/search?q={keyword}&first={page * 10 + 1}"

                self.driver.get(url)
                time.sleep(random.uniform(1, 2))

                # Extract search results
                page_competitors = self._extract_serp_results(engine)
                competitors.extend(page_competitors)

                if len(competitors) >= 15:  # Limit total results
                    break

            except Exception as e:
                logger.error(f"Error on page {page + 1} for {keyword}: {e}")
                break

        return competitors[:15]  # Return top 15 results

    def _extract_serp_results(self, engine):
        """Extract search results from SERP"""
        results = []

        try:
            if engine == 'google':
                # Google result selectors
                result_elements = self.driver.find_elements(By.CSS_SELECTOR, 'div.g')

                for element in result_elements:
                    try:
                        title_elem = element.find_element(By.CSS_SELECTOR, 'h3')
                        title = title_elem.text if title_elem else ""

                        link_elem = element.find_element(By.CSS_SELECTOR, 'a')
                        url = link_elem.get_attribute('href') if link_elem else ""

                        desc_elem = element.find_element(By.CSS_SELECTOR, '[data-sncf="1"]')
                        description = desc_elem.text if desc_elem else ""

                        if title and url and not url.startswith('https://www.google.com'):
                            results.append({
                                'title': title[:100],
                                'url': url,
                                'description': description[:200],
                                'engine': engine
                            })
                    except:
                        continue

            else:  # bing
                # Bing result selectors
                result_elements = self.driver.find_elements(By.CSS_SELECTOR, '.b_algo')

                for element in result_elements:
                    try:
                        title_elem = element.find_element(By.CSS_SELECTOR, 'h2 a')
                        title = title_elem.text if title_elem else ""
                        url = title_elem.get_attribute('href') if title_elem else ""

                        desc_elem = element.find_element(By.CSS_SELECTOR, '.b_caption p')
                        description = desc_elem.text if desc_elem else ""

                        if title and url and not url.startswith('https://www.bing.com'):
                            results.append({
                                'title': title[:100],
                                'url': url,
                                'description': description[:200],
                                'engine': engine
                            })
                    except:
                        continue

        except Exception as e:
            logger.error(f"Error extracting SERP results: {e}")

        return results

    def analyze_serp_patterns(self):
        """Analyze SERP patterns to identify keyword opportunities"""
        if not self.competitive_analysis:
            return {}

        print(f"\n📊 ANÁLISIS DE PATRONES SERP")
        print("-" * 50)

        pattern_analysis = {
            'title_patterns': Counter(),
            'description_patterns': Counter(),
            'url_patterns': Counter(),
            'common_keywords': Counter(),
            'content_gaps': [],
            'competitor_insights': {}
        }

        all_titles = []
        all_descriptions = []
        all_urls = []

        # Collect all SERP data
        for search_key, competitors in self.competitive_analysis.items():
            for competitor in competitors:
                if competitor.get('title'):
                    all_titles.append(competitor['title'].lower())
                if competitor.get('description'):
                    all_descriptions.append(competitor['description'].lower())
                if competitor.get('url'):
                    all_urls.append(competitor['url'].lower())

        # Analyze title patterns
        title_words = []
        for title in all_titles:
            words = re.findall(r'\b\w+\b', title)
            title_words.extend([word for word in words if len(word) > 3])

        pattern_analysis['title_patterns'] = Counter(title_words).most_common(20)

        # Analyze description patterns
        desc_words = []
        for desc in all_descriptions:
            words = re.findall(r'\b\w+\b', desc)
            desc_words.extend([word for word in words if len(word) > 3])

        pattern_analysis['description_patterns'] = Counter(desc_words).most_common(20)

        # Analyze common phrases
        all_text = ' '.join(all_titles + all_descriptions)
        phrases = self._extract_phrases(all_text)
        pattern_analysis['common_keywords'] = Counter(phrases).most_common(15)

        # Identify content gaps
        pattern_analysis['content_gaps'] = self._identify_content_gaps(pattern_analysis)

        self.serp_data = pattern_analysis

        print(f"✅ Análisis completado:")
        print(f"  • {len(pattern_analysis['title_patterns'])} patrones de títulos")
        print(f"  • {len(pattern_analysis['description_patterns'])} patrones de descripciones")
        print(f"  • {len(pattern_analysis['common_keywords'])} keywords principales")

        return pattern_analysis

    def _extract_phrases(self, text):
        """Extract meaningful phrases from text"""
        # Simple phrase extraction (2-3 words)
        words = re.findall(r'\b\w+\b', text.lower())
        phrases = []

        for i in range(len(words) - 1):
            if len(words[i]) > 3 and len(words[i + 1]) > 3:
                phrase = f"{words[i]} {words[i + 1]}"
                phrases.append(phrase)

        # 3-word phrases
        for i in range(len(words) - 2):
            if all(len(word) > 3 for word in words[i:i+3]):
                phrase = f"{words[i]} {words[i + 1]} {words[i + 2]}"
                phrases.append(phrase)

        return phrases

    def _identify_content_gaps(self, pattern_analysis):
        """Identify content gaps and opportunities"""
        gaps = []

        # Analyze missing common terms
        top_terms = [term for term, count in pattern_analysis['common_keywords'][:10]]

        business_activity = self.business_data.get('general_info', {}).get('actividad_principal', '').lower()
        business_location = self.business_data.get('general_info', {}).get('cobertura_geografica', '').lower()

        # Suggest location-based content
        if business_location:
            location_keywords = [
                f"{business_activity} {business_location}",
                f"mejor {business_activity} {business_location}",
                f"{business_activity} cerca de {business_location}"
            ]
            gaps.extend(location_keywords)

        # Suggest service-based content
        if business_activity:
            service_keywords = [
                f"costo {business_activity}",
                f"precio {business_activity}",
                f"como elegir {business_activity}"
            ]
            gaps.extend(service_keywords)

        return gaps[:10]

    def generate_seo_content(self):
        """Generate complete SEO content based on business data and competitive analysis"""
        print(f"\n🎨 GENERACIÓN DE CONTENIDO SEO")
        print("-" * 50)

        # Extract main keywords from competitive analysis
        main_keywords = self._extract_main_keywords()

        # Generate content for each page type
        self.generated_content = {
            'pages': {},
            'blog_posts': [],
            'faqs': {},
            'keywords_assignment': {}
        }

        page_types = ['home', 'sobre_nosotros', 'servicios', 'blog', 'contacto']

        for page_type in page_types:
            print(f"  📄 Generando contenido para: {page_type.replace('_', ' ').title()}")
            page_content = self._generate_page_content(page_type, main_keywords)
            self.generated_content['pages'][page_type] = page_content

        # Generate blog posts
        print(f"  📝 Generando artículos de blog...")
        blog_posts = self._generate_blog_posts(main_keywords)
        self.generated_content['blog_posts'] = blog_posts

        # Generate FAQs for each page
        print(f"  ❓ Generando FAQs...")
        faqs = self._generate_faqs_for_pages(main_keywords)
        self.generated_content['faqs'] = faqs

        print(f"✅ Contenido generado:")
        print(f"  • {len(self.generated_content['pages'])} páginas principales")
        print(f"  • {len(self.generated_content['blog_posts'])} artículos de blog")
        print(f"  • {len(self.generated_content['faqs'])} sets de FAQs")

        return self.generated_content

    def _extract_main_keywords(self):
        """Extract main keywords from business data and competitive analysis"""
        keywords = []

        # Business-based keywords
        business_info = self.business_data.get('general_info', {})
        activity = business_info.get('actividad_principal', '')
        location = business_info.get('cobertura_geografica', '')

        if activity:
            keywords.extend([
                activity.lower(),
                f"{activity.lower()} profesional",
                f"servicios {activity.lower()}"
            ])

        if location:
            keywords.extend([
                f"{activity.lower()} {location.lower()}",
                f"mejor {activity.lower()} {location.lower()}"
            ])

        # Add keywords from competitive analysis
        if self.serp_data.get('common_keywords'):
            comp_keywords = [kw for kw, count in self.serp_data['common_keywords'][:5]]
            keywords.extend(comp_keywords)

        return keywords[:10]

    def _generate_page_content(self, page_type, keywords):
        """Generate content for a specific page type"""
        business_name = self.business_data.get('general_info', {}).get('nombre_comercial', 'Nuestro Negocio')
        activity = self.business_data.get('general_info', {}).get('actividad_principal', 'nuestros servicios')
        location = self.business_data.get('general_info', {}).get('cobertura_geografica', '')

        # Select main keyword for this page
        main_keyword = keywords[0] if keywords else activity

        content = {
            'main_keyword': main_keyword,
            'title_seo': '',
            'meta_description': '',
            'h1': '',
            'slug': '',
            'h2_structure': [],
            'paragraphs': [],
            'cta': self.business_data.get('general_info', {}).get('cta_global', 'Contáctanos')
        }

        if page_type == 'home':
            content.update({
                'title_seo': f"{business_name} - {activity.title()} {location}",
                'meta_description': f"Servicios profesionales de {activity} en {location}. {business_name} ofrece soluciones de calidad. ¡Contáctanos hoy!",
                'h1': f"{business_name} - {activity.title()} {location}",
                'slug': '/',
                'h2_structure': [
                    f"¿Por qué elegir nuestros servicios de {main_keyword}?",
                    f"Nuestros servicios de {activity}",
                    f"Experiencia en {main_keyword} {location}",
                    f"Proceso de trabajo en {activity}",
                    f"Beneficios de trabajar con {business_name}",
                    "Testimonios de clientes satisfechos",
                    "Áreas de cobertura",
                    "Preguntas Frecuentes"
                ]
            })

        elif page_type == 'sobre_nosotros':
            content.update({
                'title_seo': f"Sobre {business_name} - Expertos en {activity.title()}",
                'meta_description': f"Conoce la historia de {business_name}, nuestro equipo y experiencia en {activity}. Más de X años brindando servicios de calidad.",
                'h1': f"Sobre {business_name} - Expertos en {activity.title()}",
                'slug': '/sobre-nosotros',
                'h2_structure': [
                    f"Historia de {business_name}",
                    f"Nuestro equipo especializado en {main_keyword}",
                    "Misión y valores",
                    f"Experiencia en {activity}",
                    "Certificaciones y reconocimientos",
                    "Nuestro compromiso con la calidad",
                    "Clientes satisfechos",
                    "Preguntas Frecuentes"
                ]
            })

        elif page_type == 'servicios':
            content.update({
                'title_seo': f"Servicios de {activity.title()} - {business_name} {location}",
                'meta_description': f"Descubre todos nuestros servicios de {activity} en {location}. Calidad, experiencia y resultados garantizados. ¡Cotiza ahora!",
                'h1': f"Servicios de {activity.title()} - {business_name}",
                'slug': '/servicios',
                'h2_structure': [
                    f"Servicios principales de {main_keyword}",
                    f"Proceso de {activity}",
                    "Garantías y calidad",
                    "Precios y cotizaciones",
                    f"Áreas de especialización en {main_keyword}",
                    "Materiales y herramientas",
                    "Tiempos de entrega",
                    "Preguntas Frecuentes"
                ]
            })

        elif page_type == 'blog':
            content.update({
                'title_seo': f"Blog - Consejos y Tips sobre {activity.title()}",
                'meta_description': f"Artículos, consejos y novedades sobre {activity}. Mantente informado con los expertos de {business_name}.",
                'h1': f"Blog - Todo sobre {activity.title()}",
                'slug': '/blog',
                'h2_structure': [
                    f"Últimos artículos sobre {main_keyword}",
                    f"Consejos profesionales de {activity}",
                    "Tendencias del sector",
                    "Casos de éxito",
                    f"Guías prácticas de {main_keyword}",
                    "Novedades del mercado",
                    "Artículos destacados",
                    "Preguntas Frecuentes"
                ]
            })

        elif page_type == 'contacto':
            content.update({
                'title_seo': f"Contacto - {business_name} {location}",
                'meta_description': f"Contacta con {business_name} para {activity} en {location}. Teléfono, WhatsApp, email. ¡Respuesta rápida garantizada!",
                'h1': f"Contacta con {business_name}",
                'slug': '/contacto',
                'h2_structure': [
                    "Información de contacto",
                    f"Solicita tu cotización de {main_keyword}",
                    "Horarios de atención",
                    "Ubicación y cobertura",
                    f"Formulario de contacto para {activity}",
                    "Redes sociales",
                    "Tiempo de respuesta",
                    "Preguntas Frecuentes"
                ]
            })

        # Ensure titles and descriptions meet SEO requirements
        content['title_seo'] = content['title_seo'][:60]
        content['meta_description'] = content['meta_description'][:160]

        return content

    def _generate_blog_posts(self, keywords):
        """Generate 4 blog post templates"""
        business_name = self.business_data.get('general_info', {}).get('nombre_comercial', 'Nuestro Negocio')
        activity = self.business_data.get('general_info', {}).get('actividad_principal', 'nuestros servicios')
        location = self.business_data.get('general_info', {}).get('cobertura_geografica', '')

        blog_posts = []

        # Blog post templates
        post_templates = [
            {
                'type': 'guide',
                'title_base': f"Guía completa de {activity}",
                'topic': f"Todo lo que necesitas saber sobre {activity}"
            },
            {
                'type': 'tips',
                'title_base': f"10 consejos para elegir {activity}",
                'topic': f"Tips profesionales para {activity}"
            },
            {
                'type': 'costs',
                'title_base': f"¿Cuánto cuesta {activity} en {location}?",
                'topic': f"Precios y factores que influyen en {activity}"
            },
            {
                'type': 'trends',
                'title_base': f"Tendencias 2024 en {activity}",
                'topic': f"Novedades y futuro de {activity}"
            }
        ]

        for i, template in enumerate(post_templates):
            main_keyword = keywords[i % len(keywords)] if keywords else activity

            post = {
                'title_seo': template['title_base'][:60],
                'meta_description': f"{template['topic']} en {location}. Consejos de expertos de {business_name}. ¡Lee más!"[:160],
                'h1': template['title_base'],
                'slug': f"/blog/{template['type']}-{activity.replace(' ', '-').lower()}",
                'main_keyword': main_keyword,
                'h2_structure': [
                    f"Introducción a {main_keyword}",
                    f"Beneficios de {main_keyword}",
                    f"Tipos de {main_keyword}",
                    f"Cómo elegir {main_keyword}",
                    f"Errores comunes en {main_keyword}",
                    f"Recomendaciones de expertos",
                    "Conclusiones",
                    "Preguntas Frecuentes"
                ],
                'type': template['type']
            }

            blog_posts.append(post)

        return blog_posts

    def _generate_faqs_for_pages(self, keywords):
        """Generate FAQs for each page"""
        business_name = self.business_data.get('general_info', {}).get('nombre_comercial', 'Nuestro Negocio')
        activity = self.business_data.get('general_info', {}).get('actividad_principal', 'nuestros servicios')
        location = self.business_data.get('general_info', {}).get('cobertura_geografica', '')

        faqs = {}

        # Common FAQ patterns
        common_faqs = [
            {
                'question': f"¿Qué incluye el servicio de {activity}?",
                'answer': f"Nuestro servicio de {activity} incluye evaluación inicial, desarrollo del trabajo, garantía y soporte post-servicio."
            },
            {
                'question': f"¿Cuánto tiempo toma el {activity}?",
                'answer': f"El tiempo de {activity} depende del alcance del proyecto. Generalmente entre 1-4 semanas."
            },
            {
                'question': f"¿Ofrecen garantía en {activity}?",
                'answer': f"Sí, todos nuestros servicios de {activity} incluyen garantía de calidad y satisfacción."
            },
            {
                'question': f"¿Atienden en {location}?",
                'answer': f"Sí, brindamos servicios de {activity} en toda la zona de {location} y alrededores."
            },
            {
                'question': f"¿Cómo solicitar una cotización de {activity}?",
                'answer': f"Puedes contactarnos por teléfono, WhatsApp o nuestro formulario web para cotizar {activity}."
            }
        ]

        # Assign FAQs to pages
        for page_type in ['home', 'sobre_nosotros', 'servicios', 'blog', 'contacto']:
            faqs[page_type] = common_faqs.copy()

        return faqs

    def generate_complete_brief_markdown(self, output_file):
        """Generate complete brief in markdown format"""
        print(f"\n📝 GENERANDO BRIEF COMPLETO")
        print("-" * 50)

        business_name = self.business_data.get('general_info', {}).get('nombre_comercial', 'Nuestro Negocio')
        activity = self.business_data.get('general_info', {}).get('actividad_principal', 'nuestros servicios')

        # Build complete brief markdown
        brief_md = f"""# Brief Completo de SEO - {business_name}
*Generado automáticamente el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*

## 📋 Datos Generales del Negocio

### Información Básica
- **Nombre comercial:** {self.business_data.get('general_info', {}).get('nombre_comercial', '')}
- **Descripción corta:** {self.business_data.get('general_info', {}).get('descripcion_corta', '')}
- **Actividad principal:** {self.business_data.get('general_info', {}).get('actividad_principal', '')}
- **Cobertura geográfica:** {self.business_data.get('general_info', {}).get('cobertura_geografica', '')}
- **Medios de pago:** {self.business_data.get('general_info', {}).get('medios_pago', '')}
- **Objetivo principal:** {self.business_data.get('general_info', {}).get('objetivo_principal', '')}
- **CTA global:** {self.business_data.get('general_info', {}).get('cta_global', '')}

### Redes Sociales
- **Instagram:** {self.business_data.get('social_media', {}).get('instagram', '')}
- **Facebook:** {self.business_data.get('social_media', {}).get('facebook', '')}
- **LinkedIn:** {self.business_data.get('social_media', {}).get('linkedin', '')}

## 🎯 Estrategia de Palabras Clave

### Keywords Principales Identificadas
"""

        # Add keyword research results
        if self.serp_data.get('common_keywords'):
            brief_md += "\n**Top Keywords de la Competencia:**\n"
            for i, (keyword, count) in enumerate(self.serp_data['common_keywords'][:10], 1):
                brief_md += f"{i}. **{keyword}** ({count} menciones)\n"

        # Add generated content for each page
        brief_md += "\n\n## 📄 Contenido SEO Generado\n\n"

        for page_type, content in self.generated_content.get('pages', {}).items():
            page_title = page_type.replace('_', ' ').title()
            brief_md += f"### {page_title}\n\n"
            brief_md += f"- **Título SEO:** {content.get('title_seo', '')}\n"
            brief_md += f"- **Meta Description:** {content.get('meta_description', '')}\n"
            brief_md += f"- **H1:** {content.get('h1', '')}\n"
            brief_md += f"- **Slug:** {content.get('slug', '')}\n"
            brief_md += f"- **Keyword Principal:** {content.get('main_keyword', '')}\n"

            brief_md += "\n**Estructura de H2:**\n"
            for i, h2 in enumerate(content.get('h2_structure', []), 1):
                brief_md += f"{i}. {h2}\n"
            brief_md += "\n"

        # Add blog posts
        brief_md += "## 📝 Artículos de Blog Sugeridos\n\n"
        for i, post in enumerate(self.generated_content.get('blog_posts', []), 1):
            brief_md += f"### Artículo {i}: {post.get('title_seo', '')}\n"
            brief_md += f"- **Meta Description:** {post.get('meta_description', '')}\n"
            brief_md += f"- **Keyword Principal:** {post.get('main_keyword', '')}\n"
            brief_md += f"- **Slug:** {post.get('slug', '')}\n\n"

        # Add competitive analysis summary
        if self.competitive_analysis:
            brief_md += "## 🔍 Resumen de Análisis Competitivo\n\n"
            brief_md += f"- **Total de competidores analizados:** {sum(len(competitors) for competitors in self.competitive_analysis.values())}\n"
            brief_md += f"- **Motores de búsqueda consultados:** Google, Bing\n"
            brief_md += f"- **Keywords investigadas:** {len(self.competitive_analysis)} términos\n"

        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(brief_md)
            print(f"✅ Brief completo guardado en: {output_file}")
            logger.info(f"Complete brief saved to: {output_file}")
        except Exception as e:
            logger.error(f"Error saving complete brief: {e}")
            print(f"❌ Error guardando brief: {e}")

    def generate_expanded_excel_report(self, output_file):
        """Generate Excel report with 4 sheets: Brief + 3 template sheets"""
        print(f"\n📊 GENERANDO REPORTE EXCEL EXPANDIDO")
        print("-" * 50)

        try:
            wb = Workbook()

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Create 4 sheets: Brief + Template structure
            self._create_business_brief_sheet(wb)  # Sheet 1: Brief with scraped data
            self._create_link_building_sheet(wb)   # Sheet 2: Link building
            self._create_blog_sheet(wb)            # Sheet 3: Blog
            self._create_seo_onpage_sheet(wb)      # Sheet 4: SEO On-Page

            # Save workbook
            wb.save(output_file)
            print(f"✅ Excel expandido guardado en: {output_file}")
            logger.info(f"Expanded Excel report saved to: {output_file}")

        except Exception as e:
            logger.error(f"Error generating expanded Excel report: {e}")
            print(f"❌ Error generando Excel: {e}")

    def generate_phase1_excel(self, output_file):
        """
        FASE 1: Genera Excel con datos base (sin recomendaciones IA)
        - Extrae datos del sitio (URL, H1, H2, FAQs existentes)
        - Deja columna 'Palabra clave principal' VACÍA (usuario la completa)
        - Deja columnas de recomendaciones IA VACÍAS (se generan en Fase 2)
        """
        print(f"\n📊 GENERANDO EXCEL FASE 1 (SIN RECOMENDACIONES IA)")
        print("-" * 50)

        try:
            wb = Workbook()

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Create sheets with Phase 1 logic (no AI recommendations)
            self._create_business_brief_sheet(wb)
            self._create_link_building_sheet(wb)
            self._create_blog_sheet_phase1(wb)      # Phase 1: No AI content
            self._create_seo_onpage_sheet_phase1(wb)  # Phase 1: No AI content

            # Save workbook
            wb.save(output_file)
            print(f"✅ Excel Fase 1 guardado en: {output_file}")
            logger.info(f"Phase 1 Excel saved to: {output_file}")

        except Exception as e:
            logger.error(f"Error generating Phase 1 Excel: {e}")
            print(f"❌ Error generando Excel Fase 1: {e}")

    def generate_phase2_excel(self, excel_file):
        """
        FASE 2: Lee Excel editado por usuario y genera recomendaciones IA
        - Lee keywords definidas por usuario en columna 7
        - Genera contenido IA solo para filas con keyword definida
        - Actualiza columnas verdes con recomendaciones
        """
        print(f"\n🤖 GENERANDO CONTENIDO IA FASE 2")
        print("-" * 50)

        try:
            if not os.path.exists(excel_file):
                logger.error(f"Excel file not found: {excel_file}")
                print(f"❌ Error: Archivo no encontrado: {excel_file}")
                return

            # Load existing Excel
            wb = load_workbook(excel_file)

            processed_total = 0
            skipped_total = 0

            # Process Blog sheet
            if "Blog" in wb.sheetnames:
                print("📝 Procesando hoja 'Blog'...")
                ws = wb["Blog"]
                processed, skipped = self._process_phase2_sheet(ws, "Blog")
                processed_total += processed
                skipped_total += skipped
                print(f"   ✅ {processed} páginas procesadas")
                print(f"   ⏭️  {skipped} páginas sin keyword (omitidas)")

            # Process SEO On-Page sheet
            if "SEO On-Page" in wb.sheetnames:
                print("📝 Procesando hoja 'SEO On-Page'...")
                ws = wb["SEO On-Page"]
                processed, skipped = self._process_phase2_sheet(ws, "SEO On-Page")
                processed_total += processed
                skipped_total += skipped
                print(f"   ✅ {processed} páginas procesadas")
                print(f"   ⏭️  {skipped} páginas sin keyword (omitidas)")

            # Save updated Excel
            wb.save(excel_file)
            print(f"\n✅ Excel Fase 2 guardado en: {excel_file}")
            print(f"📊 Total procesado: {processed_total} páginas")
            print(f"⏭️  Total omitido: {skipped_total} páginas (sin keyword)")
            logger.info(f"Phase 2 Excel saved: {excel_file} ({processed_total} processed, {skipped_total} skipped)")

        except Exception as e:
            logger.error(f"Error generating Phase 2 Excel: {e}")
            print(f"❌ Error generando Excel Fase 2: {e}")
            import traceback
            traceback.print_exc()

    def _process_phase2_sheet(self, ws, sheet_name):
        """
        Procesa una hoja en Fase 2 (Blog o SEO On-Page)
        Lee keywords del usuario y genera contenido IA
        """
        processed = 0
        skipped = 0

        # Start from row 2 (skip header)
        for row in range(2, ws.max_row + 1):
            url = ws.cell(row, 5).value  # Column 5: URL
            user_keyword = ws.cell(row, 7).value  # Column 7: Palabra clave principal

            # Skip if no keyword defined by user
            if not user_keyword or str(user_keyword).strip() == "":
                skipped += 1
                continue

            user_keyword = str(user_keyword).strip()

            # Get existing data
            h1 = ws.cell(row, 8).value or ""  # Column 8: H1
            h2 = ws.cell(row, 10).value or ""  # Column 10: H2
            existing_faqs = ws.cell(row, 12).value or ""  # Column 12: FAQs

            # Find page data for contextual generation
            page = self._find_page_by_url(url) if url else None

            # Detect niche for contextual recommendations
            niche, confidence = ('general', 0)
            if page:
                niche, confidence = self._detect_business_niche(page)
                logger.info(f"   🧠 Niche detected: '{niche}' (confidence: {confidence})")

            # Generate AI recommendations using user-defined keyword
            logger.info(f"   🔄 Generando contenido para keyword: '{user_keyword}'")

            url_recommendation = self._generate_seo_url_recommendation(url, user_keyword) if url else ""
            h1_recommendation = self._generate_seo_h1_recommendation(h1, user_keyword)
            h2_recommendation = self._generate_seo_h2_recommendation(h2, user_keyword, page)
            faqs_recommendation = self._generate_seo_faqs_recommendation(existing_faqs, user_keyword, page)

            # Update recommendation columns (green columns)
            ws.cell(row, 6, value=url_recommendation)  # Column 6: ✅ URL Optimizada
            ws.cell(row, 9, value=h1_recommendation)   # Column 9: ✅ H1 Optimizado
            ws.cell(row, 11, value=h2_recommendation)  # Column 11: ✅ H2 Optimizados
            ws.cell(row, 13, value=faqs_recommendation)  # Column 13: ✅ FAQs Optimizadas

            processed += 1

        return processed, skipped

    def _find_page_by_url(self, url):
        """Encuentra datos de página por URL en content_data"""
        if not url or not self.content_data:
            return None

        url = str(url).strip()
        for page in self.content_data:
            page_url = page.get('url', '').strip()
            if page_url == url:
                return page

        return None

    def run_enhanced_analysis(self, sitemap_path=None, url_list_path=None, output_file="analisis-mejorado.md", excel_output="analisis-mejorado.xlsx"):
        """Enhanced analysis for existing sites with content generation capabilities"""
        print(f"\n🔍 ANÁLISIS MEJORADO DE SITIO EXISTENTE")
        print("="*60)

        # Run original analysis first
        self.run_analysis(sitemap_path, url_list_path, output_file, excel_output)

        # Extract business data from existing content if possible
        if self.content_data:
            self._extract_business_data_from_content()

        # Perform competitive research based on extracted keywords
        main_keywords = [kw for kw, count in self.keywords.most_common(3)]
        if main_keywords:
            print(f"\n🔍 Realizando investigación competitiva...")
            self.search_competitors(main_keywords)
            self.analyze_serp_patterns()

        # Generate additional content recommendations
        self.generate_seo_content()

        # Generate enhanced Excel report
        self.generate_expanded_excel_report(excel_output)

        print(f"\n✅ Análisis mejorado completado:")
        print(f"  📝 Reporte markdown: {output_file}")
        print(f"  📊 Excel expandido: {excel_output}")

    def _extract_business_data_from_content(self):
        """Extract comprehensive business data from existing website content via web scraping"""
        # Initialize data structures
        if not self.business_data.get('general_info'):
            self.business_data['general_info'] = {}
        if not self.business_data.get('social_media'):
            self.business_data['social_media'] = {}
        if not self.business_data.get('contact'):
            self.business_data['contact'] = {}
        if not self.business_data.get('services'):
            self.business_data['services'] = {}

        # Collect all data from scraped pages
        all_titles = [page.get('title', '') for page in self.content_data if page.get('title')]
        all_content = [page.get('body_text', '') for page in self.content_data if page.get('body_text')]
        all_meta_descriptions = [page.get('meta_description', '') for page in self.content_data if page.get('meta_description')]
        all_h1 = []
        all_h2 = []
        for page in self.content_data:
            if page.get('h1'):
                all_h1.extend(page['h1'])
            if page.get('h2'):
                all_h2.extend(page['h2'])

        combined_text = ' '.join(all_titles + all_content + all_meta_descriptions).lower()

        # 1. Extract business name
        if not self.business_data['general_info'].get('nombre_comercial'):
            # Try from home page title first
            for page in self.content_data:
                url = page.get('url', '').lower()
                if url.endswith('/') or url.split('/')[-1] in ['index.html', 'home', '']:
                    title = page.get('title', '')
                    if title:
                        # Remove common suffixes
                        clean_name = re.split(r'[-|–—:]', title)[0].strip()
                        self.business_data['general_info']['nombre_comercial'] = clean_name
                        break

            # Fallback: use first non-generic title
            if not self.business_data['general_info'].get('nombre_comercial'):
                for title in all_titles[:3]:
                    if title and len(title.split()) <= 6 and not any(generic in title.lower() for generic in ['home', 'inicio', 'blog', 'contacto', 'aviso', 'política']):
                        self.business_data['general_info']['nombre_comercial'] = title.split('-')[0].strip()
                        break

        # 2. Extract short description (from meta descriptions or first paragraph)
        if not self.business_data['general_info'].get('descripcion_corta'):
            if all_meta_descriptions:
                # Use the longest, most descriptive meta description
                best_desc = max([desc for desc in all_meta_descriptions if len(desc) > 50],
                              key=len, default='')
                if best_desc:
                    self.business_data['general_info']['descripcion_corta'] = best_desc[:300]

            # If still no description, generate one from content
            if not self.business_data['general_info'].get('descripcion_corta'):
                self.business_data['general_info']['descripcion_corta'] = self._generate_description_from_content()

        # 3. Extract main activity (from about page, services, or most common topics)
        if not self.business_data['general_info'].get('actividad_principal'):
            # Look for "sobre nosotros" or "about" pages
            about_content = []
            for page in self.content_data:
                url = page.get('url', '').lower()
                if any(keyword in url for keyword in ['sobre', 'about', 'nosotros', 'quienes', 'empresa']):
                    about_content.append(page.get('body_text', ''))

            if about_content:
                combined_about = ' '.join(about_content)
                # Extract first meaningful sentence (simplified)
                sentences = combined_about.split('.')
                for sentence in sentences[:5]:
                    if len(sentence) > 30 and len(sentence) < 200:
                        self.business_data['general_info']['actividad_principal'] = sentence.strip() + '.'
                        break

            # If still no activity, generate from content analysis
            if not self.business_data['general_info'].get('actividad_principal'):
                self.business_data['general_info']['actividad_principal'] = self._generate_main_activity_from_content()

        # 4. Extract service zones (zonas de servicio)
        if not self.business_data['general_info'].get('zonas_servicio'):
            # Look for country/city names in content
            location_keywords = []
            spanish_cities = ['madrid', 'barcelona', 'valencia', 'sevilla', 'bilbao', 'málaga', 'valladolid', 'zaragoza']
            countries = ['españa', 'mexico', 'argentina', 'colombia', 'chile', 'perú']

            for location in spanish_cities + countries:
                if location in combined_text:
                    location_keywords.append(location.title())

            if location_keywords:
                self.business_data['general_info']['zonas_servicio'] = ', '.join(set(location_keywords[:3]))

        # 5. Extract payment methods
        if not self.business_data['general_info'].get('medios_pago'):
            payment_keywords = {
                'tarjeta': 'Tarjeta de crédito/débito',
                'paypal': 'PayPal',
                'transferencia': 'Transferencia bancaria',
                'efectivo': 'Efectivo',
                'bizum': 'Bizum',
                'stripe': 'Stripe'
            }
            found_payments = []
            for keyword, payment_name in payment_keywords.items():
                if keyword in combined_text:
                    found_payments.append(payment_name)

            if found_payments:
                self.business_data['general_info']['medios_pago'] = ', '.join(found_payments[:4])

        # 6. Extract differentiators (from unique selling points)
        if not self.business_data['general_info'].get('diferenciales'):
            differentiator_keywords = ['único', 'exclusiv', 'diferente', 'especializado', 'líder', 'mejor', 'experto', 'profesional']
            differentiators = []

            for page in self.content_data:
                text = page.get('body_text', '').lower()
                for keyword in differentiator_keywords:
                    if keyword in text:
                        # Extract sentence containing keyword
                        sentences = text.split('.')
                        for sentence in sentences:
                            if keyword in sentence and len(sentence) > 20 and len(sentence) < 150:
                                differentiators.append(sentence.strip())
                                break

            if differentiators:
                self.business_data['general_info']['diferenciales'] = list(set(differentiators[:3]))

            # If no differentiators found, generate from content analysis
            if not self.business_data['general_info'].get('diferenciales'):
                self.business_data['general_info']['diferenciales'] = self._generate_differentiators_from_content()

        # 7. Extract social media links
        social_patterns = {
            'instagram': r'instagram\.com/([a-zA-Z0-9._]+)',
            'facebook': r'facebook\.com/([a-zA-Z0-9._]+)',
            'linkedin': r'linkedin\.com/(company|in)/([a-zA-Z0-9._-]+)',
            'twitter': r'twitter\.com/([a-zA-Z0-9._]+)',
            'youtube': r'youtube\.com/(channel|c|user)/([a-zA-Z0-9._-]+)'
        }

        for page in self.content_data:
            body_text = page.get('body_text', '')
            for platform, pattern in social_patterns.items():
                if not self.business_data['social_media'].get(platform):
                    match = re.search(pattern, body_text)
                    if match:
                        if platform == 'linkedin':
                            self.business_data['social_media'][platform] = f"https://linkedin.com/{match.group(1)}/{match.group(2)}"
                        elif platform == 'youtube':
                            self.business_data['social_media'][platform] = f"https://youtube.com/{match.group(1)}/{match.group(2)}"
                        else:
                            self.business_data['social_media'][platform] = f"https://{platform}.com/{match.group(1)}"

        # 8. Extract contact information
        # Email
        if not self.business_data['contact'].get('email'):
            email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
            for page in self.content_data:
                text = page.get('body_text', '')
                email_match = re.search(email_pattern, text)
                if email_match:
                    email = email_match.group(0)
                    # Exclude common generic emails in examples
                    if not any(ex in email.lower() for ex in ['example', 'correo', 'email', 'test']):
                        self.business_data['contact']['email'] = email
                        break

        # Phone
        if not self.business_data['contact'].get('telefono'):
            phone_patterns = [
                r'\+?\d{1,3}[-.\s]?\(?\d{1,4}\)?[-.\s]?\d{1,4}[-.\s]?\d{1,9}',
                r'\d{3}[-.\s]?\d{3}[-.\s]?\d{3}'
            ]
            for page in self.content_data:
                url = page.get('url', '').lower()
                if 'contacto' in url or 'contact' in url:
                    text = page.get('body_text', '')
                    for pattern in phone_patterns:
                        phone_match = re.search(pattern, text)
                        if phone_match:
                            self.business_data['contact']['telefono'] = phone_match.group(0)
                            break
                    if self.business_data['contact'].get('telefono'):
                        break

        # 9. Extract services/products
        if not self.business_data['services'].get('lista_servicios'):
            services = []
            # Look in services/products pages
            for page in self.content_data:
                url = page.get('url', '').lower()
                if any(keyword in url for keyword in ['servicio', 'service', 'producto', 'product']):
                    h2_list = page.get('h2', [])
                    h3_list = page.get('h3', [])
                    # H2 and H3 often list services
                    services.extend([h.strip() for h in h2_list if len(h.strip()) > 5 and len(h.strip()) < 100])
                    services.extend([h.strip() for h in h3_list if len(h.strip()) > 5 and len(h.strip()) < 100])

            if services:
                self.business_data['services']['lista_servicios'] = list(set(services[:10]))

        # 10. Extract main objective/CTA
        if not self.business_data['general_info'].get('objetivo_principal'):
            cta_keywords = ['contacta', 'solicita', 'reserva', 'compra', 'suscríbe', 'regístrate', 'descarga', 'prueba']
            for page in self.content_data:
                text = page.get('body_text', '').lower()
                for keyword in cta_keywords:
                    if keyword in text:
                        # Find common CTAs
                        if 'contacta' in text or 'solicita' in text:
                            self.business_data['general_info']['objetivo_principal'] = 'Generar leads / consultas'
                            break
                        elif 'compra' in text or 'reserva' in text:
                            self.business_data['general_info']['objetivo_principal'] = 'Ventas online'
                            break
                        elif 'suscríbe' in text or 'regístrate' in text:
                            self.business_data['general_info']['objetivo_principal'] = 'Captación de registros'
                            break
                if self.business_data['general_info'].get('objetivo_principal'):
                    break

        # 11. Extract mission (misión)
        if not self.business_data['general_info'].get('mision'):
            for page in self.content_data:
                text = page.get('body_text', '')
                # Look for mission keywords
                if 'misión' in text.lower() or 'nuestra misión' in text.lower():
                    # Extract paragraph after "misión"
                    sentences = text.split('.')
                    for i, sentence in enumerate(sentences):
                        if 'misión' in sentence.lower() and i < len(sentences) - 1:
                            mission_text = sentences[i] + '.' + (sentences[i+1] if i+1 < len(sentences) else '')
                            if len(mission_text) > 30 and len(mission_text) < 300:
                                self.business_data['general_info']['mision'] = mission_text.strip()
                                break
                if self.business_data['general_info'].get('mision'):
                    break

            # If no mission found, generate from content
            if not self.business_data['general_info'].get('mision'):
                self.business_data['general_info']['mision'] = self._generate_mission_from_content()

        # 12. Extract vision (visión)
        if not self.business_data['general_info'].get('vision'):
            for page in self.content_data:
                text = page.get('body_text', '')
                if 'visión' in text.lower() or 'nuestra visión' in text.lower():
                    sentences = text.split('.')
                    for i, sentence in enumerate(sentences):
                        if 'visión' in sentence.lower() and i < len(sentences) - 1:
                            vision_text = sentences[i] + '.' + (sentences[i+1] if i+1 < len(sentences) else '')
                            if len(vision_text) > 30 and len(vision_text) < 300:
                                self.business_data['general_info']['vision'] = vision_text.strip()
                                break
                if self.business_data['general_info'].get('vision'):
                    break

            # If no vision found, generate from content
            if not self.business_data['general_info'].get('vision'):
                self.business_data['general_info']['vision'] = self._generate_vision_from_content()

        # 13. Extract values (valores)
        if not self.business_data['general_info'].get('valores'):
            values_list = []
            for page in self.content_data:
                text = page.get('body_text', '')
                if 'valores' in text.lower() or 'nuestros valores' in text.lower():
                    # Try to extract values from H2/H3 near "valores"
                    h2_list = page.get('h2', [])
                    h3_list = page.get('h3', [])

                    # Common value keywords
                    value_keywords = ['integridad', 'honestidad', 'compromiso', 'excelencia', 'innovación',
                                    'transparencia', 'responsabilidad', 'respeto', 'calidad', 'confianza']

                    for h in h2_list + h3_list:
                        if any(keyword in h.lower() for keyword in value_keywords):
                            values_list.append(h.strip())

                    if values_list:
                        self.business_data['general_info']['valores'] = ', '.join(values_list[:5])
                        break

            # If no explicit values found, generate SEO-optimized values based on content analysis
            if not values_list and self.content_data:
                values_list = self._generate_seo_values_from_content()
                if values_list:
                    self.business_data['general_info']['valores'] = ', '.join(values_list)

        logger.info(f"Extracted business data from {len(self.content_data)} scraped pages")
        logger.info(f"Business name: {self.business_data.get('general_info', {}).get('nombre_comercial', 'Not found')}")
        logger.info(f"Services found: {len(self.business_data.get('services', {}).get('lista_servicios', []))}")
        logger.info(f"Social media: {list(self.business_data.get('social_media', {}).keys())}")

    def _generate_description_from_content(self):
        """Generate SEO-optimized short description from content analysis"""
        combined_text = ' '.join([page.get('body_text', '') for page in self.content_data[:3]])[:500]
        business_name = self.business_data.get('general_info', {}).get('nombre_comercial', 'Este negocio')

        # Extract key service words from H2/H3
        service_words = []
        for page in self.content_data[:3]:
            service_words.extend(page.get('h2', [])[:2])
            service_words.extend(page.get('h3', [])[:2])

        service_context = ', '.join(service_words[:3]) if service_words else 'servicios especializados'

        return f"{business_name} ofrece {service_context.lower()} con atención personalizada, calidad garantizada y compromiso con la excelencia en cada proyecto."

    def _generate_main_activity_from_content(self):
        """Generate main activity description from content analysis"""
        # Analyze most common H2 titles and keywords
        h2_titles = []
        for page in self.content_data:
            h2_titles.extend(page.get('h2', []))

        if h2_titles:
            # Use most common H2 as activity indicator
            from collections import Counter
            common_topics = Counter(h2_titles).most_common(3)
            topic = common_topics[0][0] if common_topics else 'servicios profesionales'
        else:
            topic = 'servicios especializados'

        business_name = self.business_data.get('general_info', {}).get('nombre_comercial', 'La empresa')
        return f"{business_name} se especializa en {topic.lower()}, brindando soluciones integrales y personalizadas para satisfacer las necesidades de cada cliente."

    def _generate_differentiators_from_content(self):
        """Generate business differentiators from content analysis"""
        differentiators = []

        combined_text = ' '.join([page.get('body_text', '') for page in self.content_data]).lower()

        # Define differentiator patterns
        if 'experiencia' in combined_text or 'años' in combined_text:
            differentiators.append('Amplia experiencia en el sector con resultados comprobados')

        if 'personalizado' in combined_text or 'medida' in combined_text:
            differentiators.append('Atención personalizada y soluciones a medida para cada cliente')

        if 'calidad' in combined_text or 'excelencia' in combined_text:
            differentiators.append('Compromiso con la calidad y excelencia en cada proyecto')

        if 'equipo' in combined_text or 'profesional' in combined_text:
            differentiators.append('Equipo altamente capacitado de profesionales especializados')

        # Default if nothing found
        if not differentiators:
            differentiators = [
                'Enfoque centrado en resultados y satisfacción del cliente',
                'Innovación constante en metodologías y procesos',
                'Transparencia y comunicación fluida en cada etapa del proyecto'
            ]

        return differentiators[:3]

    def _generate_mission_from_content(self):
        """Generate mission statement from content analysis"""
        business_name = self.business_data.get('general_info', {}).get('nombre_comercial', 'Nuestra empresa')
        activity = self.business_data.get('general_info', {}).get('actividad_principal', '')

        if activity:
            core_service = activity.split('.')[0].lower().replace(business_name.lower(), '').strip()
        else:
            core_service = 'brindar servicios de excelencia'

        return f"La misión de {business_name} es {core_service}, superando las expectativas de nuestros clientes mediante la innovación, el compromiso y la calidad en cada proyecto que emprendemos."

    def _generate_vision_from_content(self):
        """Generate vision statement from content analysis"""
        business_name = self.business_data.get('general_info', {}).get('nombre_comercial', 'Nuestra organización')

        combined_text = ' '.join([page.get('body_text', '') for page in self.content_data]).lower()

        # Determine industry context
        if any(word in combined_text for word in ['tecnología', 'digital', 'software', 'web']):
            sector = 'en el sector tecnológico'
        elif any(word in combined_text for word in ['consultoría', 'asesoría', 'gestión']):
            sector = 'en consultoría empresarial'
        elif any(word in combined_text for word in ['educación', 'formación', 'cursos']):
            sector = 'en el ámbito educativo'
        else:
            sector = 'en nuestro sector'

        return f"Ser reconocidos como líderes {sector}, destacándonos por nuestra innovación, calidad de servicio y compromiso inquebrantable con el éxito de nuestros clientes."

    def _generate_seo_values_from_content(self):
        """Generate SEO-optimized business values based on content analysis"""
        values_list = []

        # Analyze content to infer values
        combined_text = ' '.join([page.get('body_text', '') for page in self.content_data]).lower()

        # Define value inference patterns based on content themes
        value_patterns = {
            'Calidad y Excelencia': ['mejor', 'calidad', 'premium', 'superior', 'destacado', 'óptimo', 'excelente'],
            'Compromiso con el Cliente': ['cliente', 'satisfacción', 'atención', 'servicio', 'apoyo', 'asesoramiento'],
            'Innovación y Tecnología': ['innovador', 'tecnología', 'moderno', 'avanzado', 'digital', 'actualizado'],
            'Experiencia y Profesionalismo': ['experiencia', 'años', 'profesional', 'experto', 'especializado', 'trayectoria'],
            'Transparencia y Confianza': ['confianza', 'transparente', 'honesto', 'seguro', 'garantía', 'certificado'],
            'Responsabilidad Social': ['sostenible', 'responsable', 'comunidad', 'social', 'medio ambiente', 'ecológico'],
            'Eficiencia y Resultados': ['eficiente', 'efectivo', 'rápido', 'resultados', 'optimizado', 'productivo']
        }

        # Score each value based on keyword frequency
        value_scores = {}
        for value, keywords in value_patterns.items():
            score = sum(combined_text.count(keyword) for keyword in keywords)
            if score > 0:
                value_scores[value] = score

        # Get top 4-5 values based on content
        if value_scores:
            sorted_values = sorted(value_scores.items(), key=lambda x: x[1], reverse=True)
            values_list = [value for value, score in sorted_values[:5]]

        # If still no values found, provide default professional values
        if not values_list:
            business_name = self.business_data.get('general_info', {}).get('nombre_comercial', '')
            activity = self.business_data.get('general_info', {}).get('actividad_principal', '')

            # Generate contextual default values
            if any(word in activity.lower() for word in ['tecnología', 'software', 'digital', 'web']):
                values_list = ['Innovación Tecnológica', 'Calidad en el Desarrollo', 'Atención Personalizada', 'Resultados Medibles']
            elif any(word in activity.lower() for word in ['consultoría', 'asesoría', 'servicios']):
                values_list = ['Experiencia Comprobada', 'Compromiso con el Cliente', 'Soluciones Personalizadas', 'Transparencia']
            elif any(word in activity.lower() for word in ['comercio', 'venta', 'tienda', 'productos']):
                values_list = ['Calidad en Productos', 'Atención al Cliente', 'Precios Competitivos', 'Confianza y Seguridad']
            else:
                values_list = ['Excelencia Profesional', 'Compromiso con la Calidad', 'Orientación al Cliente', 'Integridad y Confianza']

        return values_list

    def _detect_main_keyword_with_ai(self, page):
        """Detect main keyword using AI analysis of title, meta description, and content"""
        title = page.get('title', '')
        meta_description = page.get('meta_description', '')
        body_text = page.get('body_text', '')
        h1 = page.get('h1', [''])[0] if page.get('h1') else ''

        # Combine all text sources for analysis
        combined_text = f"{title} {meta_description} {h1} {body_text}".lower()

        # Extract word frequency (excluding common stop words)
        from collections import Counter
        import re

        # Extended stop words: Spanish, English, navigation terms, months
        stop_words = {
            # Spanish stop words
            'el', 'la', 'de', 'que', 'y', 'a', 'en', 'un', 'ser', 'se', 'no', 'haber', 'por', 'con', 'su',
            'para', 'como', 'estar', 'tener', 'le', 'lo', 'todo', 'pero', 'más', 'hacer', 'o', 'poder',
            'decir', 'este', 'ir', 'otro', 'ese', 'si', 'me', 'ya', 'ver', 'porque', 'dar', 'cuando',
            'él', 'muy', 'sin', 'vez', 'mucho', 'saber', 'qué', 'sobre', 'mi', 'alguno', 'mismo', 'yo',
            'también', 'hasta', 'año', 'dos', 'querer', 'entre', 'así', 'primero', 'desde', 'grande',
            # English stop words
            'the', 'be', 'to', 'of', 'and', 'a', 'in', 'that', 'have', 'i', 'it', 'for', 'not', 'on',
            'with', 'he', 'as', 'you', 'do', 'at', 'this', 'but', 'his', 'by', 'from', 'they', 'we',
            'say', 'her', 'she', 'or', 'an', 'will', 'my', 'one', 'all', 'would', 'there', 'their', 'is',
            'was', 'are', 'has', 'been', 'www', 'com', 'http', 'https',
            # Navigation and common WordPress terms
            'inicio', 'home', 'contacto', 'contact', 'servicios', 'services', 'nosotros', 'about',
            'empresa', 'company', 'privacidad', 'privacy', 'aviso', 'legal', 'terminos', 'terms',
            'cookies', 'politica', 'policy', 'categoria', 'category', 'archivo', 'archive', 'tag',
            'autor', 'author', 'buscar', 'search', 'pagina', 'page', 'menu', 'sidebar', 'widget',
            'footer', 'header', 'navigation', 'navegacion',
            # Months in Spanish
            'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto',
            'septiembre', 'octubre', 'noviembre', 'diciembre',
            # Months in English
            'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august',
            'september', 'october', 'november', 'december',
            # Generic filler words
            'aqui', 'alli', 'here', 'there', 'cuando', 'where', 'what', 'quien', 'como', 'donde',
            'cual', 'quien', 'cuanto', 'cuantos'
        }

        # Tokenize and count words
        words = re.findall(r'\b[a-záéíóúñ]{3,}\b', combined_text)
        filtered_words = [word for word in words if word not in stop_words and len(word) > 3]

        word_counts = Counter(filtered_words)

        # Score keywords based on position and frequency
        keyword_scores = {}

        for word, count in word_counts.most_common(20):
            score = count

            # Boost score if word appears in title (x3)
            if word in title.lower():
                score *= 3

            # Boost score if word appears in meta description (x2)
            if word in meta_description.lower():
                score *= 2

            # Boost score if word appears in H1 (x2.5)
            if word in h1.lower():
                score *= 2.5

            keyword_scores[word] = score

        # Get top keyword
        if keyword_scores:
            sorted_keywords = sorted(keyword_scores.items(), key=lambda x: x[1], reverse=True)
            main_keyword = sorted_keywords[0][0]

            # Quality guard: check if keyword is suspicious (navigation/month term)
            suspicious_terms = {
                'inicio', 'contacto', 'servicios', 'nosotros', 'home', 'contact', 'about',
                'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto',
                'septiembre', 'octubre', 'noviembre', 'diciembre', 'january', 'february',
                'march', 'april', 'june', 'july', 'august', 'september', 'october', 'november', 'december'
            }

            # If main keyword is suspicious, try next best keyword
            if main_keyword in suspicious_terms and len(sorted_keywords) > 1:
                main_keyword = sorted_keywords[1][0]

            # Check for 2-word phrases (bigrams) with the main keyword
            bigram_pattern = rf'\b\w+\s+{main_keyword}\b|\b{main_keyword}\s+\w+\b'
            bigrams = re.findall(bigram_pattern, combined_text)

            if bigrams:
                bigram_counts = Counter(bigrams)
                most_common_bigram = bigram_counts.most_common(1)[0]
                if most_common_bigram[1] >= 3:  # If bigram appears 3+ times
                    final_keyword = most_common_bigram[0].strip()
                    # Mark as "REVISAR" if still contains suspicious terms
                    if any(term in final_keyword.lower() for term in suspicious_terms):
                        return f"⚠️ REVISAR: {final_keyword}"
                    return final_keyword

            # Mark main keyword for review if suspicious
            if main_keyword in suspicious_terms:
                return f"⚠️ REVISAR: {main_keyword}"

            return main_keyword
        else:
            # Fallback: use title words
            title_words = re.findall(r'\b[a-záéíóúñ]{4,}\b', title.lower())
            title_filtered = [w for w in title_words if w not in stop_words]
            return title_filtered[0] if title_filtered else 'contenido'

    def _generate_seo_url_recommendation(self, url, keyword):
        """Generate SEO-optimized URL recommendation"""
        if not keyword:
            return url

        from urllib.parse import urlparse
        parsed = urlparse(url)
        path_parts = [p for p in parsed.path.split('/') if p]

        # Get the last meaningful part of the URL
        if path_parts:
            last_part = path_parts[-1]
            # Remove file extension if exists
            if '.' in last_part:
                last_part = last_part.rsplit('.', 1)[0]

            # Clean keyword for URL (lowercase, replace spaces with hyphens)
            clean_keyword = keyword.lower().replace(' ', '-').replace('_', '-')

            # If keyword not in URL, suggest adding it
            if clean_keyword not in last_part.lower():
                # Create SEO-friendly slug
                suggested_slug = f"{clean_keyword}-{last_part}" if last_part else clean_keyword
                suggested_url = f"{parsed.scheme}://{parsed.netloc}/{'/'.join(path_parts[:-1])}/{suggested_slug}".rstrip('/')
                return suggested_url

        return url

    def _generate_seo_h1_recommendation(self, h1, keyword):
        """Generate SEO-optimized H1 recommendation"""
        if not keyword or not h1:
            return h1 if h1 else f"Guía Completa de {keyword.title()}" if keyword else ""

        keyword_lower = keyword.lower()
        h1_lower = h1.lower()

        # If keyword already in H1, optimize structure
        if keyword_lower in h1_lower:
            # Check if it's at the beginning (best practice)
            if not h1_lower.startswith(keyword_lower):
                return f"{keyword.title()}: {h1}"
            return h1
        else:
            # Add keyword naturally to H1
            power_words = ['Guía Completa', 'Todo lo que Necesitas Saber', 'Mejores Prácticas', 'Expertos en']
            return f"{keyword.title()}: {h1} - Guía Completa"

    def _extract_real_topics_from_content(self, page, keyword):
        """Extract real topics from content using TF-IDF-like approach"""
        body_text = page.get('body_text', '')
        h2_list = page.get('h2', [])

        if not body_text:
            return []

        # Tokenize content into sentences
        sentences = re.split(r'[.!?]+', body_text)

        # Extended stop words
        stop_words = {
            'el', 'la', 'de', 'que', 'y', 'a', 'en', 'un', 'ser', 'se', 'no', 'por', 'con', 'su',
            'para', 'como', 'estar', 'tener', 'lo', 'todo', 'más', 'hacer', 'este', 'muy', 'sin',
            'the', 'be', 'to', 'of', 'and', 'in', 'that', 'have', 'it', 'for', 'not', 'on', 'with',
            'inicio', 'contacto', 'servicios', 'nosotros', 'home', 'about', 'contact'
        }

        # Extract meaningful bigrams from content
        word_pattern = r'\b[a-záéíóúñ]{4,}\b'
        bigrams = []

        for sentence in sentences[:30]:  # Analyze first 30 sentences
            words = [w.lower() for w in re.findall(word_pattern, sentence) if w.lower() not in stop_words]
            for i in range(len(words) - 1):
                bigram = f"{words[i]} {words[i+1]}"
                if len(bigram) > 8 and keyword and keyword.lower() in bigram:
                    bigrams.append(bigram)

        # Count bigram frequency
        bigram_counts = Counter(bigrams)

        # Extract topics from existing H2s
        topics_from_h2 = []
        if h2_list:
            for h2 in h2_list[:5]:
                if len(h2) > 10 and h2.lower().strip() not in ['', 'undefined']:
                    topics_from_h2.append(h2.strip())

        # Combine: prioritize existing H2s, then add discovered bigrams
        final_topics = topics_from_h2[:3]  # Keep up to 3 existing H2s

        for bigram, count in bigram_counts.most_common(5):
            if len(final_topics) < 5:
                # Convert bigram to question format
                topic = f"¿Qué es {bigram}?" if 'qué' not in bigram.lower() else bigram.title()
                if topic not in final_topics:
                    final_topics.append(topic)

        return final_topics

    def _generate_seo_h2_recommendation(self, h2_text, keyword, page=None):
        """Generate contextual, intelligent H2 recommendations based on industry and URL structure"""
        if not keyword:
            return h2_text

        # Skip if keyword is marked for review
        if keyword.startswith('⚠️ REVISAR'):
            keyword = keyword.replace('⚠️ REVISAR: ', '')

        if not page:
            # Fallback to generic if no page data
            return self._generate_generic_h2s(h2_text, keyword)

        # STEP 1: Detect business niche and URL structure
        niche, confidence = self._detect_business_niche(page)
        url_structure = self._parse_url_structure(page.get('url', ''))

        # STEP 2: Get industry knowledge bank
        industry_banks = self._get_industry_knowledge_banks()
        industry_data = industry_banks.get(niche, None)

        # STEP 3: Determine page type from URL structure
        page_type = url_structure['section']  # servicios, blog, about, etc.

        # STEP 4: Extract real topics from content
        real_topics = []
        if page:
            real_topics = self._extract_real_topics_from_content(page, keyword)

        # STEP 5: Build contextual H2s
        enhanced_h2s = []

        # Start with existing real H2s (max 2)
        if h2_text and h2_text.strip():
            h2_list = [h.strip() for h in h2_text.split(',') if h.strip()]
            enhanced_h2s.extend(h2_list[:2])

        # Add real topics from bigrams (max 2)
        for topic in real_topics[:2]:
            if topic not in enhanced_h2s:
                enhanced_h2s.append(topic)

        # STEP 6: Generate contextual H2s based on niche + page type
        if industry_data:
            h2_templates = industry_data['h2_templates']

            # Apply templates with keyword replacement
            for template in h2_templates:
                if len(enhanced_h2s) >= 5:
                    break

                # Replace {keyword} placeholder
                contextual_h2 = template.replace('{keyword}', keyword)

                # Skip if already added
                if contextual_h2 not in enhanced_h2s:
                    enhanced_h2s.append(contextual_h2)
        else:
            # Fallback to generic variations
            generic_variations = [
                f"¿Qué es {keyword} y cómo funciona?",
                f"Beneficios principales de {keyword}",
                f"Cómo elegir el mejor {keyword}",
                f"Precio y opciones de {keyword}",
                f"Guía completa sobre {keyword}"
            ]

            for variation in generic_variations:
                if len(enhanced_h2s) >= 5:
                    break
                if variation not in enhanced_h2s:
                    enhanced_h2s.append(variation)

        return ', '.join(enhanced_h2s[:5])

    def _generate_generic_h2s(self, h2_text, keyword):
        """Fallback generic H2 generation when no page data available"""
        keyword_variations = [
            f"¿Qué es {keyword} y cómo funciona?",
            f"Beneficios principales de {keyword}",
            f"Cómo elegir el mejor {keyword}",
            f"Precio y opciones de {keyword}",
            f"Guía completa sobre {keyword}"
        ]

        if not h2_text or h2_text.strip() == '':
            return ', '.join(keyword_variations[:5])

        h2_list = [h.strip() for h in h2_text.split(',') if h.strip()]
        enhanced_h2s = h2_list[:2]

        while len(enhanced_h2s) < 5:
            for suggestion in keyword_variations:
                if suggestion not in enhanced_h2s:
                    enhanced_h2s.append(suggestion)
                    break
            if len(enhanced_h2s) >= 5:
                break

        return ', '.join(enhanced_h2s[:5])

    def _generate_seo_faqs_recommendation(self, existing_faqs, keyword, page=None):
        """Generate contextual, intelligent FAQs based on detected industry"""
        if not keyword:
            return existing_faqs

        # Skip if keyword is marked for review
        if keyword.startswith('⚠️ REVISAR'):
            keyword = keyword.replace('⚠️ REVISAR: ', '')

        # STEP 1: Detect business niche if page available
        niche = 'general'
        if page:
            niche, confidence = self._detect_business_niche(page)

        # STEP 2: Get industry-specific questions
        industry_banks = self._get_industry_knowledge_banks()
        industry_data = industry_banks.get(niche, None)

        enhanced_faqs = []

        # STEP 3: Use industry-specific common questions
        if industry_data and 'common_questions' in industry_data:
            common_questions = industry_data['common_questions']

            for i, question in enumerate(common_questions[:5], 1):
                # Generate contextual answer based on niche
                answer = self._generate_contextual_answer(question, keyword, niche, industry_data)
                enhanced_faqs.append(f"P{i}: {question}\nR: {answer}")

        else:
            # Fallback to generic FAQs
            generic_faq_templates = [
                {
                    'q': f"¿Qué es {keyword} y para qué sirve?",
                    'a': f"{keyword.title()} es una solución profesional que permite optimizar procesos, mejorar resultados y aumentar la eficiencia. Se utiliza principalmente para alcanzar objetivos específicos con calidad garantizada."
                },
                {
                    'q': f"¿Cuáles son los principales beneficios de {keyword}?",
                    'a': f"Los beneficios de {keyword} incluyen: mejora en la productividad, optimización de recursos, resultados medibles, atención personalizada y soporte continuo para garantizar el éxito."
                },
                {
                    'q': f"¿Cómo elegir el mejor servicio de {keyword}?",
                    'a': f"Para elegir el mejor servicio de {keyword}, considera: experiencia comprobada, casos de éxito, opiniones de clientes, metodología transparente y atención personalizada que se adapte a tus necesidades específicas."
                },
                {
                    'q': f"¿Cuánto tiempo se necesita para ver resultados con {keyword}?",
                    'a': f"Los resultados con {keyword} pueden variar según el proyecto, pero generalmente se observan mejoras significativas en las primeras semanas. El compromiso y la implementación correcta son clave para el éxito."
                },
                {
                    'q': f"¿Dónde puedo encontrar servicios profesionales de {keyword}?",
                    'a': f"Puedes encontrar servicios profesionales de {keyword} contactando con especialistas certificados que ofrezcan garantías, casos de éxito demostrados y atención personalizada para tu proyecto."
                }
            ]

            for i, faq in enumerate(generic_faq_templates, 1):
                enhanced_faqs.append(f"P{i}: {faq['q']}\nR: {faq['a']}")

        return "\n\n".join(enhanced_faqs)

    def _generate_contextual_answer(self, question, keyword, niche, industry_data):
        """Generate contextual answer based on industry and question"""
        question_lower = question.lower()

        # Detect question type and provide specific answers
        if 'cuánto' in question_lower and ('cuesta' in question_lower or 'cobra' in question_lower or 'precio' in question_lower):
            # Price-related question
            if niche == 'fisioterapia':
                return f"El precio de {keyword} varía según el tipo de tratamiento y número de sesiones necesarias. Una sesión individual suele costar entre 35-60€. Muchas clínicas ofrecen bonos de 5 o 10 sesiones con descuento."
            elif niche == 'dentista':
                return f"El coste de {keyword} depende del tratamiento específico. Una consulta de diagnóstico ronda los 30-50€. Tratamientos más complejos pueden financiarse en cómodas cuotas mensuales."
            elif niche == 'abogado':
                return f"Los honorarios para {keyword} pueden variar según la complejidad del caso. Algunos despachos ofrecen primera consulta gratuita y trabajan con tarifas planas o por éxito."
            elif niche == 'restaurante':
                return f"El precio del {keyword} en nuestro restaurante es competitivo y ajustado a la calidad ofrecida. Consulta nuestra carta actualizada o menú del día para conocer precios específicos."
            elif niche == 'gimnasio':
                return f"La cuota mensual para {keyword} varía según el plan elegido. Ofrecemos desde 25€/mes en plan básico hasta 50€/mes con acceso completo y clases ilimitadas."
            else:
                return f"El precio de {keyword} depende de varios factores como la duración, complejidad y servicios incluidos. Contacta para recibir un presupuesto personalizado sin compromiso."

        elif 'cuánto' in question_lower and ('dura' in question_lower or 'tiempo' in question_lower or 'tarda' in question_lower):
            # Duration-related question
            if niche == 'fisioterapia':
                return f"La duración del tratamiento de {keyword} depende del tipo y severidad de la lesión. Las sesiones suelen durar 45-60 minutos. La recuperación completa puede tomar de 2-3 semanas a varios meses según el caso."
            elif niche == 'dentista':
                return f"El tiempo necesario para {keyword} varía según el procedimiento. Tratamientos simples pueden completarse en una visita, mientras que otros como ortodoncia pueden durar de 6 meses a 2 años."
            elif niche == 'abogado':
                return f"La duración del proceso de {keyword} depende de la complejidad del caso y la vía judicial. Casos simples pueden resolverse en 3-6 meses, mientras que procedimientos complejos pueden extenderse hasta 1-2 años."
            else:
                return f"El tiempo necesario para {keyword} varía según cada situación particular. Te proporcionaremos un cronograma detallado tras la evaluación inicial de tu caso específico."

        elif 'sesiones' in question_lower or 'visitas' in question_lower or 'citas' in question_lower:
            # Sessions-related question
            if niche == 'fisioterapia':
                return f"El número de sesiones de {keyword} necesarias depende de la condición tratada. Problemas agudos pueden mejorar en 3-5 sesiones, mientras que condiciones crónicas pueden requerir 10-15 sesiones de tratamiento."
            elif niche == 'psicologia':
                return f"La cantidad de sesiones de {keyword} varía según cada persona. Algunas situaciones se resuelven en 5-10 sesiones, mientras que procesos más profundos pueden requerir terapia a medio-largo plazo."
            elif niche == 'estetica':
                return f"El número de sesiones de {keyword} necesarias depende del tipo de piel y objetivo deseado. Generalmente se recomiendan entre 6-10 sesiones para obtener resultados óptimos y duraderos."
            else:
                return f"La cantidad de sesiones necesarias para {keyword} se determina tras una evaluación inicial. Cada caso es único y requiere un plan personalizado para garantizar los mejores resultados."

        elif 'horario' in question_lower or 'abierto' in question_lower or 'cerrado' in question_lower:
            # Schedule-related question
            return f"Nuestro horario de atención para {keyword} es de lunes a viernes de 9:00 a 20:00h y sábados de 9:00 a 14:00h. También ofrecemos citas con horarios flexibles para adaptarnos a tu disponibilidad."

        elif 'urgencias' in question_lower or 'emergencias' in question_lower:
            # Emergency-related question
            if niche in ['veterinario', 'fontanero', 'electricista']:
                return f"Sí, ofrecemos servicio de urgencias 24 horas para {keyword}. Contamos con un equipo disponible para atender cualquier emergencia de forma inmediata. Llama al teléfono de urgencias."
            else:
                return f"Para casos urgentes de {keyword}, ofrecemos citas prioritarias el mismo día o al día siguiente. Contacta lo antes posible para evaluar tu situación y darte la atención necesaria."

        elif 'primera' in question_lower and ('consulta' in question_lower or 'cita' in question_lower or 'visita' in question_lower):
            # First visit question
            return f"Sí, ofrecemos primera consulta de {keyword} gratuita o con precio especial. Es una oportunidad para conocernos, evaluar tu situación y diseñar un plan personalizado sin compromiso."

        elif 'online' in question_lower or 'distancia' in question_lower or 'remoto' in question_lower:
            # Online service question
            return f"Sí, ofrecemos servicios de {keyword} online mediante videollamada. Es una opción cómoda y efectiva que te permite recibir atención profesional desde la comodidad de tu hogar."

        elif 'financiación' in question_lower or 'financiar' in question_lower or 'cuotas' in question_lower:
            # Financing question
            return f"Sí, para {keyword} ofrecemos opciones de financiación flexible en cómodas cuotas mensuales sin intereses. Queremos que el precio no sea un impedimento para recibir la atención que necesitas."

        else:
            # Generic answer with industry context
            services = industry_data.get('services', [])
            services_text = ', '.join(services[:3]) if services else keyword
            return f"En {keyword}, ofrecemos servicios profesionales especializados incluyendo {services_text}, entre otros. Contamos con experiencia comprobada y metodología efectiva para garantizar tu satisfacción."

    def _extract_faqs_from_content(self, content_text):
        """Extract existing FAQs from page content"""
        import re
        faqs_found = []

        if not content_text:
            return []

        # Look for FAQ patterns in content
        # Pattern 1: Question followed by answer (common FAQ structure)
        question_patterns = [
            r'¿[^?]+\?',  # Spanish questions
            r'What [^?]+\?',  # English questions starting with What
            r'How [^?]+\?',   # English questions starting with How
            r'Why [^?]+\?',   # English questions starting with Why
            r'When [^?]+\?',  # English questions starting with When
            r'Where [^?]+\?', # English questions starting with Where
        ]

        for pattern in question_patterns:
            questions = re.findall(pattern, content_text, re.IGNORECASE)
            for question in questions[:10]:  # Limit to first 10 questions found
                # Try to find answer after question (next 200 chars)
                question_pos = content_text.find(question)
                if question_pos != -1:
                    answer_text = content_text[question_pos + len(question):question_pos + len(question) + 300]
                    # Clean answer: get sentences until we hit another question or end
                    sentences = answer_text.split('.')
                    answer = ''
                    for sentence in sentences[:3]:
                        if '¿' not in sentence and '?' not in sentence or sentence.count('?') == 0:
                            answer += sentence.strip() + '. '
                        else:
                            break

                    if len(answer.strip()) > 20:
                        faqs_found.append({
                            'question': question.strip(),
                            'answer': answer.strip()[:250]
                        })

        return faqs_found[:5]  # Return max 5 FAQs

    def _generate_faqs_for_content(self, title, keyword, content_text):
        """Generate 5 FAQs with answers - first tries to extract from content, then generates"""
        faqs_list = []

        # STEP 1: Try to extract existing FAQs from content
        if content_text:
            existing_faqs = self._extract_faqs_from_content(content_text)
            if existing_faqs:
                for i, faq in enumerate(existing_faqs, 1):
                    faq_entry = f"P{i}: {faq['question']}\nR: {faq['answer']}"
                    faqs_list.append(faq_entry)

        # STEP 2: If no FAQs found or not enough, generate AI-based FAQs
        if len(faqs_list) < 5 and keyword:
            remaining_count = 5 - len(faqs_list)
            faq_templates = [
                f"¿Qué es {keyword}?",
                f"¿Cómo funciona {keyword}?",
                f"¿Cuáles son los beneficios de {keyword}?",
                f"¿Por qué elegir {keyword}?",
                f"¿Dónde puedo obtener {keyword}?",
                f"¿Cuánto cuesta {keyword}?",
                f"¿Cuándo es recomendable usar {keyword}?",
                f"¿Quién puede usar {keyword}?",
            ]

            # Generate answers for remaining FAQs
            for i, question in enumerate(faq_templates[:remaining_count], len(faqs_list) + 1):
                answer = ""

                # Try to extract answer from content
                if content_text and len(content_text) > 100:
                    # Search for keyword context in content
                    keyword_lower = keyword.lower()
                    if keyword_lower in content_text.lower():
                        # Find paragraph containing keyword
                        paragraphs = content_text.split('\n')
                        for para in paragraphs:
                            if keyword_lower in para.lower() and len(para) > 50:
                                sentences = para.split('.')[:2]
                                answer = '. '.join([s.strip() for s in sentences if len(s.strip()) > 20])[:200]
                                if answer:
                                    break

                # Generate generic answer if no content-based answer found
                if not answer:
                    generic_answers = {
                        "¿Qué es": f"{keyword.title()} es un servicio/producto que ofrece soluciones especializadas para satisfacer necesidades específicas de nuestros clientes.",
                        "¿Cómo funciona": f"{keyword.title()} funciona mediante un proceso estructurado que garantiza resultados efectivos y satisfactorios para todos nuestros usuarios.",
                        "¿Cuáles son los beneficios": f"Los beneficios de {keyword} incluyen mejora en la eficiencia, resultados comprobados, y atención personalizada a cada cliente.",
                        "¿Por qué elegir": f"Elegir {keyword} significa optar por calidad, experiencia y compromiso con la excelencia en cada servicio que ofrecemos.",
                        "¿Dónde puedo obtener": f"Puedes obtener {keyword} contactando directamente con nosotros a través de nuestro sitio web o medios de contacto disponibles.",
                        "¿Cuánto cuesta": f"El costo de {keyword} varía según las necesidades específicas. Contacta con nuestro equipo para obtener una cotización personalizada.",
                        "¿Cuándo es recomendable": f"Es recomendable usar {keyword} cuando se busca optimizar procesos y obtener resultados profesionales de alta calidad.",
                        "¿Quién puede usar": f"{keyword.title()} está diseñado para todas las personas y empresas que buscan soluciones profesionales y efectivas.",
                    }

                    for key, gen_answer in generic_answers.items():
                        if key in question:
                            answer = gen_answer
                            break

                    if not answer:
                        answer = f"Para más información sobre {keyword}, te recomendamos contactar con nuestro equipo especializado que podrá asesorarte de forma personalizada."

                # Format FAQ entry
                faq_entry = f"P{i}: {question}\nR: {answer}"
                faqs_list.append(faq_entry)

        # Combine all FAQs with separator
        return "\n\n".join(faqs_list) if faqs_list else ""

    def _create_link_building_sheet(self, wb):
        """Create Link Building sheet with internal and external links"""
        ws = wb.create_sheet("Link building")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = ["Tipo de Enlace", "Página Origen", "URL Destino", "Dominio", "Texto Ancla", "Follow/NoFollow", "Estado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        row = 2

        # Extract internal and external links from scraped content
        internal_links = []
        external_links = []
        base_domain = None

        if self.content_data:
            # Determine base domain from first URL
            first_url = self.content_data[0].get('url', '')
            if first_url:
                parsed = urlparse(first_url)
                base_domain = parsed.netloc

            # Process all pages to extract links
            for page in self.content_data:
                page_url = page.get('url', '')
                body_text = page.get('body_text', '')

                # Extract links from page (simplified - looks for href patterns)
                import re
                link_pattern = r'https?://[^\s<>"\']+|www\.[^\s<>"\']+'
                found_links = re.findall(link_pattern, body_text)

                for link in found_links:
                    if not link.startswith('http'):
                        link = 'https://' + link

                    parsed_link = urlparse(link)
                    link_domain = parsed_link.netloc

                    if base_domain and link_domain == base_domain:
                        internal_links.append({
                            'origin': page_url,
                            'destination': link,
                            'domain': link_domain
                        })
                    elif link_domain:
                        external_links.append({
                            'origin': page_url,
                            'destination': link,
                            'domain': link_domain
                        })

        # Remove duplicates
        seen_internal = set()
        unique_internal = []
        for link in internal_links:
            key = (link['destination'], link['origin'])
            if key not in seen_internal:
                seen_internal.add(key)
                unique_internal.append(link)

        seen_external = set()
        unique_external = []
        for link in external_links:
            key = (link['destination'], link['origin'])
            if key not in seen_external:
                seen_external.add(key)
                unique_external.append(link)

        # SECTION: Internal Links
        section_cell = ws.cell(row=row, column=1, value="ENLACES INTERNOS")
        section_cell.font = Font(bold=True)
        section_cell.fill = section_fill
        section_cell.border = border
        for col in range(2, 8):
            ws.cell(row=row, column=col).fill = section_fill
            ws.cell(row=row, column=col).border = border
        row += 1

        for link in unique_internal[:50]:  # Limit to 50 internal links
            ws.cell(row=row, column=1, value="Interno").border = border
            ws.cell(row=row, column=2, value=link['origin']).border = border
            ws.cell(row=row, column=3, value=link['destination']).border = border
            ws.cell(row=row, column=4, value=link['domain']).border = border
            ws.cell(row=row, column=5, value="").border = border
            ws.cell(row=row, column=6, value="Follow").border = border
            ws.cell(row=row, column=7, value="Activo").border = border
            row += 1

        # SECTION: External Links
        row += 1
        section_cell = ws.cell(row=row, column=1, value="ENLACES EXTERNOS")
        section_cell.font = Font(bold=True)
        section_cell.fill = section_fill
        section_cell.border = border
        for col in range(2, 8):
            ws.cell(row=row, column=col).fill = section_fill
            ws.cell(row=row, column=col).border = border
        row += 1

        for link in unique_external[:50]:  # Limit to 50 external links
            ws.cell(row=row, column=1, value="Externo").border = border
            ws.cell(row=row, column=2, value=link['origin']).border = border
            ws.cell(row=row, column=3, value=link['destination']).border = border
            ws.cell(row=row, column=4, value=link['domain']).border = border
            ws.cell(row=row, column=5, value="").border = border
            ws.cell(row=row, column=6, value="Follow").border = border
            ws.cell(row=row, column=7, value="Verificar").border = border
            row += 1

        # SECTION: Oportunidades de Link Building (from competitive analysis)
        if self.competitive_analysis:
            row += 1
            section_cell = ws.cell(row=row, column=1, value="OPORTUNIDADES DE BACKLINKS")
            section_cell.font = Font(bold=True)
            section_cell.fill = section_fill
            section_cell.border = border
            for col in range(2, 8):
                ws.cell(row=row, column=col).fill = section_fill
                ws.cell(row=row, column=col).border = border
            row += 1

            for search_key, competitors in self.competitive_analysis.items():
                keyword = search_key.split('_', 1)[1] if '_' in search_key else search_key

                for competitor in competitors[:5]:  # Top 5 competitors per keyword
                    url = competitor.get('url', '')
                    if url:
                        domain = urlparse(url).netloc
                        ws.cell(row=row, column=1, value="Oportunidad").border = border
                        ws.cell(row=row, column=2, value="").border = border
                        ws.cell(row=row, column=3, value=url).border = border
                        ws.cell(row=row, column=4, value=domain).border = border
                        ws.cell(row=row, column=5, value=keyword).border = border
                        ws.cell(row=row, column=6, value="").border = border
                        ws.cell(row=row, column=7, value="Pendiente").border = border
                        row += 1

        self.auto_adjust_columns(ws, max_width=60)

    def _create_blog_sheet(self, wb):
        """Create Blog sheet with SEO recommendations"""
        ws = wb.create_sheet("Blog")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        recommendation_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers with recommendations columns
        headers = [
            "Fecha", "Tipo", "Reponsable", "Estado",
            "URL", "✅ URL Optimizada (Recomendación SEO)",
            "Palabra clave principal",
            "H1", "✅ H1 Optimizado (Recomendación SEO)",
            "H2", "✅ H2 Optimizados (Recomendación SEO)",
            "FAQs (Preguntas Frecuentes)", "✅ FAQs Optimizadas (Recomendación SEO)"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            # Green background for recommendation columns
            if "✅" in header or "Recomendación" in header:
                cell.fill = recommendation_fill
            else:
                cell.fill = header_fill
            cell.border = border

        # Populate with blog articles from content_data
        row = 2
        if self.content_data:
            for page in self.content_data:
                url = page.get('url', '')
                # Identify blog posts (URLs containing 'blog', 'articulo', or date patterns)
                if any(indicator in url.lower() for indicator in ['blog', 'articulo', '/20']):
                    title = page.get('title', '')
                    h1 = page.get('h1', [''])[0] if page.get('h1') else title
                    h2_list = page.get('h2', [])
                    h2 = ', '.join(h2_list[:3]) if h2_list else ''

                    # Use AI to detect main keyword from title, meta description, and content
                    main_keyword = self._detect_main_keyword_with_ai(page)

                    # Generate FAQs for this blog post
                    faqs = self._generate_faqs_for_content(title, main_keyword, page.get('body_text', ''))

                    # Generate SEO recommendations
                    url_recommendation = self._generate_seo_url_recommendation(url, main_keyword)
                    h1_recommendation = self._generate_seo_h1_recommendation(h1, main_keyword)
                    h2_recommendation = self._generate_seo_h2_recommendation(h2, main_keyword, page)
                    faqs_recommendation = self._generate_seo_faqs_recommendation(faqs, main_keyword, page)

                    # Write data
                    ws.cell(row=row, column=1, value=datetime.now().strftime('%Y-%m-%d')).border = border
                    ws.cell(row=row, column=2, value="Artículo").border = border
                    ws.cell(row=row, column=3, value="").border = border
                    ws.cell(row=row, column=4, value="Publicado").border = border
                    ws.cell(row=row, column=5, value=url).border = border
                    cell = ws.cell(row=row, column=6, value=url_recommendation)
                    cell.border = border
                    cell.fill = recommendation_fill
                    ws.cell(row=row, column=7, value=main_keyword).border = border
                    ws.cell(row=row, column=8, value=h1).border = border
                    cell = ws.cell(row=row, column=9, value=h1_recommendation)
                    cell.border = border
                    cell.fill = recommendation_fill
                    ws.cell(row=row, column=10, value=h2).border = border
                    cell = ws.cell(row=row, column=11, value=h2_recommendation)
                    cell.border = border
                    cell.fill = recommendation_fill
                    ws.cell(row=row, column=12, value=faqs).border = border
                    cell = ws.cell(row=row, column=13, value=faqs_recommendation)
                    cell.border = border
                    cell.fill = recommendation_fill
                    row += 1

        # Add generated blog content if available
        if self.generated_content.get('blog_articles'):
            for article in self.generated_content['blog_articles']:
                title = article.get('title', '')
                keyword = article.get('keyword', '')

                # Generate FAQs and recommendations
                faqs = self._generate_faqs_for_content(title, keyword, '')
                h1_recommendation = self._generate_seo_h1_recommendation(title, keyword)
                h2_recommendation = self._generate_seo_h2_recommendation('', keyword, None)
                faqs_recommendation = self._generate_seo_faqs_recommendation(faqs, keyword, None)

                ws.cell(row=row, column=1, value=datetime.now().strftime('%Y-%m-%d')).border = border
                ws.cell(row=row, column=2, value="Artículo").border = border
                ws.cell(row=row, column=3, value="").border = border
                ws.cell(row=row, column=4, value="Planificado").border = border
                ws.cell(row=row, column=5, value="").border = border
                ws.cell(row=row, column=6, value="").border = border
                ws.cell(row=row, column=7, value=keyword).border = border
                ws.cell(row=row, column=8, value=title).border = border
                cell = ws.cell(row=row, column=9, value=h1_recommendation)
                cell.border = border
                cell.fill = recommendation_fill
                ws.cell(row=row, column=10, value="").border = border
                cell = ws.cell(row=row, column=11, value=h2_recommendation)
                cell.border = border
                cell.fill = recommendation_fill
                ws.cell(row=row, column=12, value=faqs).border = border
                cell = ws.cell(row=row, column=13, value=faqs_recommendation)
                cell.border = border
                cell.fill = recommendation_fill
                row += 1

        self.auto_adjust_columns(ws, max_width=80)

    def _create_seo_onpage_sheet(self, wb):
        """Create SEO On-Page sheet with SEO recommendations"""
        ws = wb.create_sheet("SEO On-Page")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        recommendation_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers with recommendations columns
        headers = [
            "Fecha", "Tipo", "Reponsable", "Estado",
            "URL", "✅ URL Optimizada (Recomendación SEO)",
            "Palabra clave principal",
            "H1", "✅ H1 Optimizado (Recomendación SEO)",
            "H2", "✅ H2 Optimizados (Recomendación SEO)",
            "FAQs (Preguntas Frecuentes)", "✅ FAQs Optimizadas (Recomendación SEO)"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            # Green background for recommendation columns
            if "✅" in header or "Recomendación" in header:
                cell.fill = recommendation_fill
            else:
                cell.fill = header_fill
            cell.border = border

        # Populate with all pages from content_data
        row = 2
        if self.content_data:
            for page in self.content_data:
                url = page.get('url', '')
                title = page.get('title', '')
                h1 = page.get('h1', [''])[0] if page.get('h1') else title
                h2_list = page.get('h2', [])
                h2 = ', '.join(h2_list[:3]) if h2_list else ''

                # Determine page type
                page_type = "Página"
                if any(indicator in url.lower() for indicator in ['blog', 'articulo', '/20']):
                    continue  # Skip blog posts (they go in Blog sheet)

                # Use AI to detect main keyword from title, meta description, and content
                main_keyword = self._detect_main_keyword_with_ai(page)

                # Generate FAQs for this page
                faqs = self._generate_faqs_for_content(title, main_keyword, page.get('body_text', ''))

                # Generate SEO recommendations
                url_recommendation = self._generate_seo_url_recommendation(url, main_keyword)
                h1_recommendation = self._generate_seo_h1_recommendation(h1, main_keyword)
                h2_recommendation = self._generate_seo_h2_recommendation(h2, main_keyword, page)
                faqs_recommendation = self._generate_seo_faqs_recommendation(faqs, main_keyword, page)

                # Write data
                ws.cell(row=row, column=1, value=datetime.now().strftime('%Y-%m-%d')).border = border
                ws.cell(row=row, column=2, value=page_type).border = border
                ws.cell(row=row, column=3, value="").border = border
                ws.cell(row=row, column=4, value="Publicado").border = border
                ws.cell(row=row, column=5, value=url).border = border
                cell = ws.cell(row=row, column=6, value=url_recommendation)
                cell.border = border
                cell.fill = recommendation_fill
                ws.cell(row=row, column=7, value=main_keyword).border = border
                ws.cell(row=row, column=8, value=h1).border = border
                cell = ws.cell(row=row, column=9, value=h1_recommendation)
                cell.border = border
                cell.fill = recommendation_fill
                ws.cell(row=row, column=10, value=h2).border = border
                cell = ws.cell(row=row, column=11, value=h2_recommendation)
                cell.border = border
                cell.fill = recommendation_fill
                ws.cell(row=row, column=12, value=faqs).border = border
                cell = ws.cell(row=row, column=13, value=faqs_recommendation)
                cell.border = border
                cell.fill = recommendation_fill
                row += 1

        # Add generated pages if available
        if self.generated_content.get('pages'):
            for page_type, content in self.generated_content['pages'].items():
                page_name = page_type.replace('_', ' ').title()
                main_keyword = content.get('main_keyword', '')

                # Generate FAQs and recommendations
                faqs = self._generate_faqs_for_content(page_name, main_keyword, '')
                h1_recommendation = self._generate_seo_h1_recommendation(page_name, main_keyword)
                h2_recommendation = self._generate_seo_h2_recommendation('', main_keyword, None)
                faqs_recommendation = self._generate_seo_faqs_recommendation(faqs, main_keyword, None)

                ws.cell(row=row, column=1, value=datetime.now().strftime('%Y-%m-%d')).border = border
                ws.cell(row=row, column=2, value="Página").border = border
                ws.cell(row=row, column=3, value="").border = border
                ws.cell(row=row, column=4, value="Planificado").border = border
                ws.cell(row=row, column=5, value="").border = border
                ws.cell(row=row, column=6, value="").border = border
                ws.cell(row=row, column=7, value=main_keyword).border = border
                ws.cell(row=row, column=8, value=page_name).border = border
                cell = ws.cell(row=row, column=9, value=h1_recommendation)
                cell.border = border
                cell.fill = recommendation_fill
                ws.cell(row=row, column=10, value="").border = border
                cell = ws.cell(row=row, column=11, value=h2_recommendation)
                cell.border = border
                cell.fill = recommendation_fill
                ws.cell(row=row, column=12, value=faqs).border = border
                cell = ws.cell(row=row, column=13, value=faqs_recommendation)
                cell.border = border
                cell.fill = recommendation_fill
                row += 1

        self.auto_adjust_columns(ws, max_width=80)

    def _create_blog_sheet_phase1(self, wb):
        """
        FASE 1: Create Blog sheet WITHOUT AI recommendations
        Only extracts and displays existing data from scraped pages
        """
        ws = wb.create_sheet("Blog")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        recommendation_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = [
            "Fecha", "Tipo", "Reponsable", "Estado",
            "URL", "✅ URL Optimizada (Recomendación SEO)",
            "Palabra clave principal",
            "H1", "✅ H1 Optimizado (Recomendación SEO)",
            "H2", "✅ H2 Optimizados (Recomendación SEO)",
            "FAQs (Preguntas Frecuentes)", "✅ FAQs Optimizadas (Recomendación SEO)"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            if "✅" in header or "Recomendación" in header:
                cell.fill = recommendation_fill
            else:
                cell.fill = header_fill
            cell.border = border

        # Populate with scraped blog data ONLY (NO AI recommendations)
        row = 2
        if self.content_data:
            for page in self.content_data:
                url = page.get('url', '')
                # Identify blog posts
                if any(indicator in url.lower() for indicator in ['blog', 'articulo', '/20']):
                    title = page.get('title', '')
                    h1 = page.get('h1', [''])[0] if page.get('h1') else title
                    h2_list = page.get('h2', [])
                    h2 = ', '.join(h2_list[:3]) if h2_list else ''

                    # Extract existing FAQs if present
                    faqs = self._extract_existing_faqs_from_page(page)

                    # Write data (NO keyword detection, NO AI recommendations)
                    ws.cell(row=row, column=1, value=datetime.now().strftime('%Y-%m-%d')).border = border
                    ws.cell(row=row, column=2, value="Artículo").border = border
                    ws.cell(row=row, column=3, value="").border = border
                    ws.cell(row=row, column=4, value="Publicado").border = border
                    ws.cell(row=row, column=5, value=url).border = border

                    # Empty recommendation columns (Phase 2 will fill these)
                    cell = ws.cell(row=row, column=6, value="")
                    cell.border = border
                    cell.fill = recommendation_fill

                    # Empty keyword column (USER will fill this)
                    ws.cell(row=row, column=7, value="").border = border

                    ws.cell(row=row, column=8, value=h1).border = border

                    cell = ws.cell(row=row, column=9, value="")
                    cell.border = border
                    cell.fill = recommendation_fill

                    ws.cell(row=row, column=10, value=h2).border = border

                    cell = ws.cell(row=row, column=11, value="")
                    cell.border = border
                    cell.fill = recommendation_fill

                    ws.cell(row=row, column=12, value=faqs).border = border

                    cell = ws.cell(row=row, column=13, value="")
                    cell.border = border
                    cell.fill = recommendation_fill

                    row += 1

        self.auto_adjust_columns(ws, max_width=80)

    def _create_seo_onpage_sheet_phase1(self, wb):
        """
        FASE 1: Create SEO On-Page sheet WITHOUT AI recommendations
        Only extracts and displays existing data from scraped pages
        """
        ws = wb.create_sheet("SEO On-Page")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        recommendation_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = [
            "Fecha", "Tipo", "Reponsable", "Estado",
            "URL", "✅ URL Optimizada (Recomendación SEO)",
            "Palabra clave principal",
            "H1", "✅ H1 Optimizado (Recomendación SEO)",
            "H2", "✅ H2 Optimizados (Recomendación SEO)",
            "FAQs (Preguntas Frecuentes)", "✅ FAQs Optimizadas (Recomendación SEO)"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            if "✅" in header or "Recomendación" in header:
                cell.fill = recommendation_fill
            else:
                cell.fill = header_fill
            cell.border = border

        # Populate with scraped pages ONLY (NO AI recommendations)
        row = 2
        if self.content_data:
            for page in self.content_data:
                url = page.get('url', '')

                # Skip blog posts (they go in Blog sheet)
                if any(indicator in url.lower() for indicator in ['blog', 'articulo', '/20']):
                    continue

                title = page.get('title', '')
                h1 = page.get('h1', [''])[0] if page.get('h1') else title
                h2_list = page.get('h2', [])
                h2 = ', '.join(h2_list[:3]) if h2_list else ''

                # Extract existing FAQs if present
                faqs = self._extract_existing_faqs_from_page(page)

                page_type = "Página"

                # Write data (NO keyword detection, NO AI recommendations)
                ws.cell(row=row, column=1, value=datetime.now().strftime('%Y-%m-%d')).border = border
                ws.cell(row=row, column=2, value=page_type).border = border
                ws.cell(row=row, column=3, value="").border = border
                ws.cell(row=row, column=4, value="Publicado").border = border
                ws.cell(row=row, column=5, value=url).border = border

                # Empty recommendation columns (Phase 2 will fill these)
                cell = ws.cell(row=row, column=6, value="")
                cell.border = border
                cell.fill = recommendation_fill

                # Empty keyword column (USER will fill this)
                ws.cell(row=row, column=7, value="").border = border

                ws.cell(row=row, column=8, value=h1).border = border

                cell = ws.cell(row=row, column=9, value="")
                cell.border = border
                cell.fill = recommendation_fill

                ws.cell(row=row, column=10, value=h2).border = border

                cell = ws.cell(row=row, column=11, value="")
                cell.border = border
                cell.fill = recommendation_fill

                ws.cell(row=row, column=12, value=faqs).border = border

                cell = ws.cell(row=row, column=13, value="")
                cell.border = border
                cell.fill = recommendation_fill

                row += 1

        self.auto_adjust_columns(ws, max_width=80)

    def _extract_existing_faqs_from_page(self, page):
        """Extract existing FAQs from page if present"""
        body_text = page.get('body_text', '')

        # Simple FAQ detection patterns
        faq_patterns = [
            r'¿[^?]+\?[^\n]+',  # Spanish questions with answers
            r'Q:[^\n]+A:[^\n]+',  # Q&A format
        ]

        faqs = []
        for pattern in faq_patterns:
            matches = re.findall(pattern, body_text)
            faqs.extend(matches[:3])  # Limit to first 3

        return '\n'.join(faqs) if faqs else ''

    def _create_business_brief_sheet(self, wb):
        """Create Business Brief sheet with comprehensive scraped data"""
        ws = wb.create_sheet("Brief del Negocio")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = ["Campo", "Valor"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        row = 2

        # SECCIÓN: Información General
        section_cell = ws.cell(row=row, column=1, value="INFORMACIÓN GENERAL")
        section_cell.font = Font(bold=True)
        section_cell.fill = section_fill
        section_cell.border = border
        for col in range(2, 3):
            ws.cell(row=row, column=col).fill = section_fill
            ws.cell(row=row, column=col).border = border
        row += 1

        general_fields = [
            ("Nombre comercial", self.business_data.get('general_info', {}).get('nombre_comercial', '')),
            ("Descripción corta", self.business_data.get('general_info', {}).get('descripcion_corta', '')),
            ("Actividad principal", self.business_data.get('general_info', {}).get('actividad_principal', '')),
            ("Zonas de servicio", self.business_data.get('general_info', {}).get('zonas_servicio', '')),
            ("Medios de pago", self.business_data.get('general_info', {}).get('medios_pago', '')),
            ("Objetivo principal", self.business_data.get('general_info', {}).get('objetivo_principal', '')),
            ("Misión", self.business_data.get('general_info', {}).get('mision', '')),
            ("Visión", self.business_data.get('general_info', {}).get('vision', '')),
            ("Valores", self.business_data.get('general_info', {}).get('valores', '')),
        ]

        # Handle differentiators separately (could be list)
        differentiators = self.business_data.get('general_info', {}).get('diferenciales', [])
        if isinstance(differentiators, list):
            diff_text = '; '.join(differentiators) if differentiators else ''
        else:
            diff_text = differentiators
        general_fields.append(("Diferenciales", diff_text))

        for field, value in general_fields:
            ws.cell(row=row, column=1, value=field).border = border
            ws.cell(row=row, column=2, value=value).border = border
            row += 1

        # SECCIÓN: Redes Sociales
        row += 1
        section_cell = ws.cell(row=row, column=1, value="REDES SOCIALES")
        section_cell.font = Font(bold=True)
        section_cell.fill = section_fill
        section_cell.border = border
        for col in range(2, 3):
            ws.cell(row=row, column=col).fill = section_fill
            ws.cell(row=row, column=col).border = border
        row += 1

        social_fields = [
            ("Instagram", self.business_data.get('social_media', {}).get('instagram', '')),
            ("Facebook", self.business_data.get('social_media', {}).get('facebook', '')),
            ("LinkedIn", self.business_data.get('social_media', {}).get('linkedin', '')),
            ("Twitter", self.business_data.get('social_media', {}).get('twitter', '')),
            ("YouTube", self.business_data.get('social_media', {}).get('youtube', '')),
        ]

        for field, value in social_fields:
            ws.cell(row=row, column=1, value=field).border = border
            ws.cell(row=row, column=2, value=value).border = border
            row += 1

        # SECCIÓN: Contacto
        row += 1
        section_cell = ws.cell(row=row, column=1, value="CONTACTO")
        section_cell.font = Font(bold=True)
        section_cell.fill = section_fill
        section_cell.border = border
        for col in range(2, 3):
            ws.cell(row=row, column=col).fill = section_fill
            ws.cell(row=row, column=col).border = border
        row += 1

        contact_fields = [
            ("Email", self.business_data.get('contact', {}).get('email', '')),
            ("Teléfono", self.business_data.get('contact', {}).get('telefono', '')),
            ("Dirección", self.business_data.get('contact', {}).get('direccion', '')),
        ]

        for field, value in contact_fields:
            ws.cell(row=row, column=1, value=field).border = border
            ws.cell(row=row, column=2, value=value).border = border
            row += 1

        # SECCIÓN: Servicios/Productos
        row += 1
        section_cell = ws.cell(row=row, column=1, value="SERVICIOS/PRODUCTOS")
        section_cell.font = Font(bold=True)
        section_cell.fill = section_fill
        section_cell.border = border
        for col in range(2, 3):
            ws.cell(row=row, column=col).fill = section_fill
            ws.cell(row=row, column=col).border = border
        row += 1

        services = self.business_data.get('services', {}).get('lista_servicios', [])
        if services:
            for i, service in enumerate(services[:10], 1):
                ws.cell(row=row, column=1, value=f"Servicio/Producto {i}").border = border
                ws.cell(row=row, column=2, value=service).border = border
                row += 1
        else:
            ws.cell(row=row, column=1, value="Lista de servicios").border = border
            ws.cell(row=row, column=2, value="").border = border
            row += 1

        self.auto_adjust_columns(ws, max_width=80)

    def _create_competitive_analysis_sheet(self, wb):
        """Create Competitive Analysis sheet (Sheet 2)"""
        ws = wb.create_sheet("Análisis Competitivo")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = ["Keyword", "Motor", "Posición", "URL Competidor", "Título", "Descripción"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Competitive data
        row = 2
        for search_key, competitors in self.competitive_analysis.items():
            keyword = search_key.split('_', 1)[1] if '_' in search_key else search_key
            engine = search_key.split('_')[0] if '_' in search_key else 'unknown'

            for pos, competitor in enumerate(competitors, 1):
                ws.cell(row=row, column=1, value=keyword).border = border
                ws.cell(row=row, column=2, value=engine.upper()).border = border
                ws.cell(row=row, column=3, value=pos).border = border
                ws.cell(row=row, column=4, value=competitor.get('url', '')).border = border
                ws.cell(row=row, column=5, value=competitor.get('title', '')).border = border
                ws.cell(row=row, column=6, value=competitor.get('description', '')).border = border
                row += 1

        self.auto_adjust_columns(ws, max_width=80)

    def _create_keywords_assignment_sheet(self, wb):
        """Create Keywords Assignment sheet (Sheet 3)"""
        ws = wb.create_sheet("Keywords por Página")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = ["Página", "Keyword Principal", "Keywords Secundarias", "Volumen Est.", "Dificultad", "Oportunidad"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Page keyword assignments
        row = 2
        for page_type, content in self.generated_content.get('pages', {}).items():
            page_name = page_type.replace('_', ' ').title()
            main_keyword = content.get('main_keyword', '')

            # Get secondary keywords from SERP analysis
            secondary_keywords = []
            if self.serp_data.get('common_keywords'):
                secondary_keywords = [kw for kw, count in self.serp_data['common_keywords'][:3] if kw != main_keyword]

            ws.cell(row=row, column=1, value=page_name).border = border
            ws.cell(row=row, column=2, value=main_keyword).border = border
            ws.cell(row=row, column=3, value=", ".join(secondary_keywords[:3])).border = border
            ws.cell(row=row, column=4, value="Medio").border = border  # Placeholder
            ws.cell(row=row, column=5, value="Media").border = border  # Placeholder
            ws.cell(row=row, column=6, value="Alta").border = border   # Placeholder
            row += 1

        self.auto_adjust_columns(ws, max_width=60)

    def _create_generated_content_sheet(self, wb):
        """Create Generated Content sheet (Sheet 4)"""
        ws = wb.create_sheet("Contenido SEO Generado")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = ["Página", "Título SEO", "Long. Título", "Meta Description", "Long. Meta", "H1", "Slug"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Generated content data
        row = 2
        for page_type, content in self.generated_content.get('pages', {}).items():
            page_name = page_type.replace('_', ' ').title()
            title_seo = content.get('title_seo', '')
            meta_desc = content.get('meta_description', '')

            ws.cell(row=row, column=1, value=page_name).border = border
            ws.cell(row=row, column=2, value=title_seo).border = border
            ws.cell(row=row, column=3, value=len(title_seo)).border = border
            ws.cell(row=row, column=4, value=meta_desc).border = border
            ws.cell(row=row, column=5, value=len(meta_desc)).border = border
            ws.cell(row=row, column=6, value=content.get('h1', '')).border = border
            ws.cell(row=row, column=7, value=content.get('slug', '')).border = border

            # Color coding
            title_len = len(title_seo)
            if title_len == 0 or title_len > 60:
                ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif title_len < 30:
                ws.cell(row=row, column=3).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            meta_len = len(meta_desc)
            if meta_len == 0 or meta_len > 160:
                ws.cell(row=row, column=5).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif meta_len < 120:
                ws.cell(row=row, column=5).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            row += 1

        self.auto_adjust_columns(ws, max_width=80)

    def _create_content_structure_sheet(self, wb):
        """Create Content Structure sheet (Sheet 5)"""
        ws = wb.create_sheet("Estructura de Contenido")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = ["Página", "Sección (H2)", "Orden", "Keyword Focus", "CTA Incluido"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Content structure data
        row = 2
        for page_type, content in self.generated_content.get('pages', {}).items():
            page_name = page_type.replace('_', ' ').title()
            h2_structure = content.get('h2_structure', [])
            main_keyword = content.get('main_keyword', '')

            for order, h2 in enumerate(h2_structure, 1):
                ws.cell(row=row, column=1, value=page_name).border = border
                ws.cell(row=row, column=2, value=h2).border = border
                ws.cell(row=row, column=3, value=order).border = border

                # Check if H2 contains main keyword
                keyword_included = "✅" if main_keyword.lower() in h2.lower() else "❌"
                ws.cell(row=row, column=4, value=keyword_included).border = border

                # Check if it's FAQ section (always has CTA)
                cta_included = "✅" if "frecuentes" in h2.lower() else "⚠️"
                ws.cell(row=row, column=5, value=cta_included).border = border

                row += 1

        self.auto_adjust_columns(ws, max_width=70)

    def _create_blog_content_sheet(self, wb):
        """Create Blog Content sheet (Sheet 6)"""
        ws = wb.create_sheet("Contenido de Blog")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = ["#", "Tipo", "Título SEO", "Meta Description", "Keyword Principal", "Slug", "Estado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Blog posts data
        row = 2
        for i, post in enumerate(self.generated_content.get('blog_posts', []), 1):
            ws.cell(row=row, column=1, value=i).border = border
            ws.cell(row=row, column=2, value=post.get('type', '').title()).border = border
            ws.cell(row=row, column=3, value=post.get('title_seo', '')).border = border
            ws.cell(row=row, column=4, value=post.get('meta_description', '')).border = border
            ws.cell(row=row, column=5, value=post.get('main_keyword', '')).border = border
            ws.cell(row=row, column=6, value=post.get('slug', '')).border = border
            ws.cell(row=row, column=7, value="Pendiente").border = border
            row += 1

        self.auto_adjust_columns(ws, max_width=80)

    def _create_faqs_sheet(self, wb):
        """Create FAQs sheet (Sheet 7)"""
        ws = wb.create_sheet("FAQs Generadas")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = ["Página", "Pregunta (H3)", "Respuesta", "Keyword Incluida", "Optimizada"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # FAQs data
        row = 2
        for page_type, faqs in self.generated_content.get('faqs', {}).items():
            page_name = page_type.replace('_', ' ').title()

            for faq in faqs:
                ws.cell(row=row, column=1, value=page_name).border = border
                ws.cell(row=row, column=2, value=faq.get('question', '')).border = border
                ws.cell(row=row, column=3, value=faq.get('answer', '')).border = border

                # Check if answer contains business activity
                activity = self.business_data.get('general_info', {}).get('actividad_principal', '')
                keyword_included = "✅" if activity.lower() in faq.get('answer', '').lower() else "❌"
                ws.cell(row=row, column=4, value=keyword_included).border = border
                ws.cell(row=row, column=5, value="✅").border = border

                row += 1

        self.auto_adjust_columns(ws, max_width=80)

    def _create_url_tracking_sheet(self, wb):
        """Create URL Tracking sheet (Sheet 8)"""
        ws = wb.create_sheet("Seguimiento de URLs")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = ["URL/Página", "Estado Actual", "Prioridad", "Contenido Generado", "Implementado", "Fecha Límite", "Notas"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # URL tracking data
        row = 2
        for page_type, content in self.generated_content.get('pages', {}).items():
            page_name = page_type.replace('_', ' ').title()

            ws.cell(row=row, column=1, value=f"{page_name} ({content.get('slug', '')})").border = border
            ws.cell(row=row, column=2, value="Nuevo").border = border

            # Set priority based on page type
            priority = "Alta" if page_type in ['home', 'servicios', 'contacto'] else "Media"
            ws.cell(row=row, column=3, value=priority).border = border
            ws.cell(row=row, column=4, value="✅").border = border
            ws.cell(row=row, column=5, value="❌").border = border
            ws.cell(row=row, column=6, value="").border = border
            ws.cell(row=row, column=7, value="").border = border

            # Color code priority
            if priority == "Alta":
                ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif priority == "Media":
                ws.cell(row=row, column=3).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            row += 1

        self.auto_adjust_columns(ws, max_width=60)

    def _create_original_analysis_sheets(self, wb):
        """Create the original 6 analysis sheets with current data"""
        # Use existing generate_excel_report logic but only create sheets
        try:
            # Create individual sheets using existing methods
            keyword_analysis = {'top_keywords': self.keywords.most_common(50), 'top_phrases': []}
            recommendations = self.generate_seo_recommendations(self.content_data, keyword_analysis)

            # Create sheets manually without saving
            self._add_summary_sheet(wb, keyword_analysis, recommendations)
            self._add_pages_analysis_sheet(wb)
            self._add_keywords_sheet(wb, keyword_analysis)
            self._add_phrases_sheet(wb, keyword_analysis)
            self._add_recommendations_sheet(wb, recommendations)
            self._add_technical_issues_sheet(wb)

        except Exception as e:
            logger.warning(f"Could not create original analysis sheets: {e}")

    def _add_summary_sheet(self, wb, keyword_analysis, recommendations):
        """Add summary sheet to workbook"""
        ws = wb.create_sheet("Resumen Ejecutivo Original")
        # Add summary data similar to original implementation
        # This is a simplified version
        pass

    def _add_pages_analysis_sheet(self, wb):
        """Add pages analysis sheet"""
        pass  # Simplified for now

    def _add_keywords_sheet(self, wb, keyword_analysis):
        """Add keywords sheet"""
        pass  # Simplified for now

    def _add_phrases_sheet(self, wb, keyword_analysis):
        """Add phrases sheet"""
        pass  # Simplified for now

    def _add_recommendations_sheet(self, wb, recommendations):
        """Add recommendations sheet"""
        pass  # Simplified for now

    def _add_technical_issues_sheet(self, wb):
        """Add technical issues sheet"""
        pass  # Simplified for now

    def parse_sitemap(self, sitemap_path):
        """Parse sitemap.xml and extract URLs, handling both sitemap index and regular sitemaps"""
        urls = []
        try:
            if sitemap_path.startswith('http'):
                response = self.session.get(sitemap_path, verify=False)
                response.raise_for_status()
                root = ET.fromstring(response.content)
            else:
                tree = ET.parse(sitemap_path)
                root = tree.getroot()

            # Handle namespace
            namespace = {'ns': 'http://www.sitemaps.org/schemas/sitemap/0.9'}

            # Check if this is a sitemap index (contains <sitemap> elements)
            sitemap_elements = root.findall('.//ns:sitemap', namespace)
            if not sitemap_elements:
                sitemap_elements = root.findall('.//sitemap')

            if sitemap_elements:
                logger.info(f"Found sitemap index with {len(sitemap_elements)} sitemaps")
                # This is a sitemap index, parse each individual sitemap
                for sitemap_element in sitemap_elements:
                    loc_element = sitemap_element.find('ns:loc', namespace)
                    if loc_element is None:
                        loc_element = sitemap_element.find('loc')

                    if loc_element is not None:
                        sitemap_url = loc_element.text.strip()
                        logger.info(f"Parsing individual sitemap: {sitemap_url}")

                        # Recursively parse each sitemap
                        sitemap_urls = self.parse_individual_sitemap(sitemap_url)
                        urls.extend(sitemap_urls)

                        # Add small delay between sitemap requests
                        if self.delay > 0:
                            time.sleep(0.2)
            else:
                # This is a regular sitemap, parse URLs directly
                urls = self.parse_individual_sitemap(sitemap_path)

        except Exception as e:
            logger.error(f"Error parsing sitemap: {e}")

        # Filter out excluded URLs
        original_count = len(urls)
        filtered_urls = [url for url in urls if not self._should_exclude_url(url)]
        excluded_count = original_count - len(filtered_urls)

        if excluded_count > 0:
            logger.info(f"🔍 Filtered out {excluded_count} non-content URLs (WordPress/Elementor/taxonomies)")

        logger.info(f"Found {len(filtered_urls)} valid content URLs from sitemap(s)")
        return filtered_urls

    def parse_individual_sitemap(self, sitemap_path):
        """Parse an individual sitemap and extract URLs"""
        urls = []
        try:
            if sitemap_path.startswith('http'):
                response = self.session.get(sitemap_path, timeout=30, verify=False)
                response.raise_for_status()
                root = ET.fromstring(response.content)
            else:
                tree = ET.parse(sitemap_path)
                root = tree.getroot()

            # Handle namespace
            namespace = {'ns': 'http://www.sitemaps.org/schemas/sitemap/0.9'}

            # Extract URLs from this sitemap
            for url_element in root.findall('.//ns:url', namespace):
                loc = url_element.find('ns:loc', namespace)
                if loc is not None:
                    urls.append(loc.text.strip())

            # Fallback without namespace
            if not urls:
                for url_element in root.findall('.//url'):
                    loc = url_element.find('loc')
                    if loc is not None:
                        urls.append(loc.text.strip())

        except Exception as e:
            logger.error(f"Error parsing individual sitemap {sitemap_path}: {e}")

        return urls

    def load_url_list(self, url_list_path):
        """Load URLs from text file"""
        urls = []
        try:
            with open(url_list_path, 'r', encoding='utf-8') as f:
                for line in f:
                    url = line.strip()
                    if url and not url.startswith('#'):
                        urls.append(url)
        except Exception as e:
            logger.error(f"Error loading URL list: {e}")
            
        logger.info(f"Loaded {len(urls)} URLs from list")
        return urls

    def _get_domain_from_url(self, url):
        """Extract domain from URL for cache organization"""
        from urllib.parse import urlparse
        parsed = urlparse(url)
        domain = parsed.netloc.replace('www.', '')
        # Clean domain for filesystem
        domain_clean = domain.replace(':', '_').replace('/', '_')
        return domain_clean

    def _setup_domain_cache(self, first_url):
        """Setup cache directory for the current domain"""
        domain = self._get_domain_from_url(first_url)
        self.current_domain = domain
        self.domain_cache_dir = self.cache_dir / domain
        self.domain_cache_dir.mkdir(exist_ok=True, parents=True)
        logger.info(f"📁 Cache directory: .seo_cache/{domain}/")

    def _get_domain_cache_metadata_file(self):
        """Get metadata file path for current domain cache"""
        if not self.domain_cache_dir:
            return None
        return self.domain_cache_dir / "_metadata.json"

    def _needs_rescraping(self):
        """Check if domain needs re-scraping (>= 60 minutes since last scrape)"""
        metadata_file = self._get_domain_cache_metadata_file()
        if not metadata_file or not metadata_file.exists():
            return True  # No cache = needs scraping

        try:
            import json
            with open(metadata_file, 'r', encoding='utf-8') as f:
                metadata = json.load(f)
                last_scrape = datetime.fromisoformat(metadata.get('last_scrape', ''))
                age_minutes = (datetime.now() - last_scrape).total_seconds() / 60

                # OBLIGATORIO re-scrapear cada 60 minutos
                return age_minutes >= self.cache_rescrape_minutes
        except Exception as e:
            logger.warning(f"Failed to read cache metadata: {e}")
            return True  # Error = needs scraping

    def _should_delete_domain_cache(self):
        """Check if domain cache should be deleted (> 24 hours)"""
        metadata_file = self._get_domain_cache_metadata_file()
        if not metadata_file or not metadata_file.exists():
            return False

        try:
            import json
            with open(metadata_file, 'r', encoding='utf-8') as f:
                metadata = json.load(f)
                last_scrape = datetime.fromisoformat(metadata.get('last_scrape', ''))
                age_hours = (datetime.now() - last_scrape).total_seconds() / 3600
                return age_hours >= self.cache_cleanup_hours
        except Exception:
            return False

    def _save_domain_cache_metadata(self, url_count):
        """Save metadata for domain cache"""
        metadata_file = self._get_domain_cache_metadata_file()
        if not metadata_file:
            return

        try:
            import json
            metadata = {
                'domain': self.current_domain,
                'last_scrape': datetime.now().isoformat(),
                'url_count': url_count,
                'cache_version': '2.0'
            }
            with open(metadata_file, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, indent=2)
            logger.info(f"💾 Saved cache metadata for {self.current_domain}")
        except Exception as e:
            logger.warning(f"Failed to save cache metadata: {e}")

    def _get_cache_filename(self, url):
        """Generate cache filename from URL within domain directory"""
        import hashlib
        url_hash = hashlib.md5(url.encode()).hexdigest()
        if self.domain_cache_dir:
            return self.domain_cache_dir / f"{url_hash}.json"
        return self.cache_dir / f"{url_hash}.json"

    def _is_cache_valid(self, cache_file):
        """Check if cache file exists and is less than 60 minutes old for scraping"""
        if not cache_file.exists():
            return False

        file_time = datetime.fromtimestamp(cache_file.stat().st_mtime)
        current_time = datetime.now()
        age_minutes = (current_time - file_time).total_seconds() / 60

        return age_minutes < self.cache_duration_minutes

    def _should_delete_cache(self, cache_file):
        """Check if cache file should be deleted (older than 24 hours)"""
        if not cache_file.exists():
            return False

        file_time = datetime.fromtimestamp(cache_file.stat().st_mtime)
        current_time = datetime.now()
        age_hours = (current_time - file_time).total_seconds() / 3600

        return age_hours >= self.cache_cleanup_hours

    def _save_to_cache(self, url, content_data):
        """Save scraped content to cache"""
        import json
        cache_file = self._get_cache_filename(url)

        try:
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'url': url,
                    'timestamp': datetime.now().isoformat(),
                    'data': content_data
                }, f, ensure_ascii=False, indent=2)
            logger.info(f"💾 Cached data for: {url}")
        except Exception as e:
            logger.warning(f"Failed to cache data for {url}: {e}")

    def _load_from_cache(self, url):
        """Load content from cache if valid"""
        import json
        cache_file = self._get_cache_filename(url)

        if not self._is_cache_valid(cache_file):
            return None

        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                cached = json.load(f)
                cached_time = datetime.fromisoformat(cached['timestamp'])
                age_minutes = (datetime.now() - cached_time).total_seconds() / 60
                logger.info(f"📦 Using cached data for: {url} (age: {age_minutes:.1f} minutes)")
                return cached['data']
        except Exception as e:
            logger.warning(f"Failed to load cache for {url}: {e}")
            return None

    def _get_cache_stats(self):
        """Get cache statistics by domain"""
        domain_dirs = [d for d in self.cache_dir.iterdir() if d.is_dir()]

        stats = {
            'total_domains': len(domain_dirs),
            'valid_domains': 0,
            'old_domains': 0,
            'expired_domains': 0
        }

        for domain_dir in domain_dirs:
            # Temporarily set domain cache dir to read metadata
            temp_domain_cache = self.domain_cache_dir
            self.domain_cache_dir = domain_dir

            metadata_file = domain_dir / "_metadata.json"
            if not metadata_file.exists():
                continue

            try:
                import json
                with open(metadata_file, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                    last_scrape = datetime.fromisoformat(metadata.get('last_scrape', ''))
                    age_minutes = (datetime.now() - last_scrape).total_seconds() / 60
                    age_hours = age_minutes / 60

                    if age_minutes < self.cache_rescrape_minutes:
                        stats['valid_domains'] += 1
                    elif age_hours < self.cache_cleanup_hours:
                        stats['old_domains'] += 1
                    else:
                        stats['expired_domains'] += 1
            except:
                pass

            # Restore domain cache dir
            self.domain_cache_dir = temp_domain_cache

        return stats

    def _clear_expired_cache(self):
        """Remove domain caches older than 24 hours"""
        import shutil
        domain_dirs = [d for d in self.cache_dir.iterdir() if d.is_dir()]
        removed = 0
        kept_old = 0

        for domain_dir in domain_dirs:
            metadata_file = domain_dir / "_metadata.json"
            if not metadata_file.exists():
                continue

            try:
                import json
                with open(metadata_file, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                    last_scrape = datetime.fromisoformat(metadata.get('last_scrape', ''))
                    age_hours = (datetime.now() - last_scrape).total_seconds() / 3600
                    age_minutes = age_hours * 60

                    if age_hours >= self.cache_cleanup_hours:
                        # Delete domain cache older than 24 hours
                        shutil.rmtree(domain_dir)
                        removed += 1
                        logger.info(f"🗑️  Deleted cache for {metadata.get('domain')} (age: {age_hours:.1f}h)")
                    elif age_minutes >= self.cache_duration_minutes:
                        # Keep domain cache between 60 min and 24 hours
                        kept_old += 1
            except Exception as e:
                logger.warning(f"Failed to process cache for {domain_dir.name}: {e}")

        if removed > 0:
            logger.info(f"🗑️  Removed {removed} domain caches older than 24 hours")
        if kept_old > 0:
            logger.info(f"📦 Keeping {kept_old} old domain caches (1h - 24h) for potential reuse")

    def _normalize_content_structure(self, content_info):
        """Normalize content structure to flatten headings for easier access"""
        if 'headings' in content_info:
            headings = content_info.get('headings', {})
            # Add flattened heading access
            for i in range(1, 7):
                key = f'h{i}'
                if key in headings:
                    content_info[key] = headings[key]
        return content_info

    def _extract_main_content_wp(self, soup):
        """Extract main content from WordPress site using common selectors"""
        # Try WordPress and Elementor content selectors (in priority order)
        main_content_selectors = [
            'article .entry-content',
            '.entry-content',
            'article .post-content',
            '.post-content',
            '.elementor-widget-theme-post-content',
            '.elementor-post__excerpt',
            'article',
            'main',
            '.content',
            '#content',
            '.main-content',
            '#main-content',
            '.page-content',
            '#page-content'
        ]

        for selector in main_content_selectors:
            main_content = soup.select_one(selector)
            if main_content:
                logger.debug(f"✅ Main content found using selector: {selector}")
                return main_content

        # Fallback: return None (will use full soup)
        logger.debug("⚠️ Main content selector not found, using full page")
        return None

    def extract_content(self, url):
        """Extract content from a single URL with cache support"""
        # Try to load from cache first
        cached_content = self._load_from_cache(url)
        if cached_content is not None:
            return cached_content

        # If not in cache or expired, scrape fresh data
        try:
            logger.info(f"🌐 Scraping fresh data from: {url}")
            response = self.session.get(url, timeout=30, verify=False)
            response.raise_for_status()

            soup = BeautifulSoup(response.content, 'html.parser')

            # Remove script and style elements
            for script in soup(["script", "style"]):
                script.decompose()

            # Remove navigation, sidebar, footer (common noise)
            for noise in soup.select('nav, header, footer, aside, .sidebar, .menu, .navigation, .footer, .header'):
                noise.decompose()

            # Extract key elements
            title = soup.find('title')
            title_text = title.get_text().strip() if title else ""

            meta_desc = soup.find('meta', attrs={'name': 'description'})
            meta_description = meta_desc.get('content', '').strip() if meta_desc else ""

            # Extract main content using WordPress common selectors
            main_content = self._extract_main_content_wp(soup)

            # Extract headings from main content only (if found), otherwise from full page
            content_source = main_content if main_content else soup
            headings = {}
            for i in range(1, 7):
                h_tags = content_source.find_all(f'h{i}')
                headings[f'h{i}'] = [h.get_text().strip() for h in h_tags if h.get_text().strip()]

            # Extract body text from main content
            if main_content:
                body_text = main_content.get_text()
            else:
                body_text = soup.get_text()
            clean_text = re.sub(r'\s+', ' ', body_text).strip()
            
            # Extract links
            links = []
            for link in soup.find_all('a', href=True):
                href = link['href']
                text = link.get_text().strip()
                if href and text:
                    links.append({'href': href, 'text': text})
            
            # Word count
            word_count = len(clean_text.split())
            
            content_info = {
                'url': url,
                'title': title_text,
                'meta_description': meta_description,
                'headings': headings,
                'body_text': clean_text[:5000],  # Truncate for storage
                'word_count': word_count,
                'links': links[:50],  # Limit links
                'status': response.status_code
            }

            # Normalize headings structure for easier access
            content_info = self._normalize_content_structure(content_info)

            # Save to cache for future use
            self._save_to_cache(url, content_info)

            # Log H2 count for debugging
            h2_count = len(content_info.get('h2', []))
            logger.info(f"✅ Extracted content from: {url} (H1: {len(content_info.get('h1', []))}, H2: {h2_count}, H3: {len(content_info.get('h3', []))})")
            return content_info
            
        except Exception as e:
            logger.error(f"Error extracting content from {url}: {e}")
            return {
                'url': url,
                'error': str(e),
                'status': 'failed'
            }

    def analyze_keywords(self, content_data):
        """Analyze keywords and topics from content"""
        all_text = ""
        
        for content in content_data:
            if content.get('status') == 'failed':
                continue
                
            text_parts = [
                content.get('title', ''),
                content.get('meta_description', ''),
                content.get('body_text', '')
            ]
            
            # Add heading text
            headings = content.get('headings', {})
            for h_level, h_list in headings.items():
                text_parts.extend(h_list)
            
            all_text += " " + " ".join(text_parts)
        
        # Clean and tokenize
        clean_text = re.sub(r'[^\w\s]', ' ', all_text.lower())
        words = clean_text.split()
        
        # Filter common stop words (basic list)
        stop_words = {
            'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 
            'by', 'is', 'are', 'was', 'were', 'be', 'been', 'have', 'has', 'had', 'do', 'does', 
            'did', 'will', 'would', 'could', 'should', 'may', 'might', 'must', 'can', 'this', 
            'that', 'these', 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they', 'me', 'him', 
            'her', 'us', 'them', 'my', 'your', 'his', 'her', 'its', 'our', 'their', 'el', 'la', 
            'los', 'las', 'un', 'una', 'y', 'o', 'pero', 'en', 'de', 'con', 'por', 'para', 'que'
        }
        
        filtered_words = [word for word in words if len(word) > 2 and word not in stop_words]
        
        # Count keywords
        self.keywords = Counter(filtered_words)
        
        # Extract common phrases (2-3 words)
        phrases = []
        for i in range(len(filtered_words) - 1):
            phrase = " ".join(filtered_words[i:i+2])
            if len(phrase) > 6:
                phrases.append(phrase)
        
        phrase_counter = Counter(phrases)
        
        logger.info(f"Analyzed {len(filtered_words)} words, found {len(self.keywords)} unique keywords")
        
        return {
            'top_keywords': self.keywords.most_common(50),
            'top_phrases': phrase_counter.most_common(25)
        }

    def generate_seo_recommendations(self, content_data, keyword_analysis):
        """Generate SEO recommendations based on analysis"""
        recommendations = []
        
        # Analyze content quality
        pages_analyzed = len([c for c in content_data if c.get('status') != 'failed'])
        pages_failed = len([c for c in content_data if c.get('status') == 'failed'])
        
        recommendations.append({
            'category': 'Análisis General',
            'priority': 'Alta',
            'action': f'Se analizaron {pages_analyzed} páginas exitosamente y {pages_failed} fallaron',
            'details': 'Revisar URLs que fallaron en el análisis'
        })
        
        # Title analysis
        no_title = len([c for c in content_data if not c.get('title')])
        short_titles = len([c for c in content_data if c.get('title') and len(c['title']) < 30])
        long_titles = len([c for c in content_data if c.get('title') and len(c['title']) > 60])
        
        if no_title > 0:
            recommendations.append({
                'category': 'Títulos',
                'priority': 'Alta',
                'action': f'Agregar títulos a {no_title} páginas que no tienen title tag',
                'details': 'Cada página debe tener un título único y descriptivo'
            })
            
        if short_titles > 0:
            recommendations.append({
                'category': 'Títulos',
                'priority': 'Media',
                'action': f'Optimizar {short_titles} títulos que son muy cortos (menos de 30 caracteres)',
                'details': 'Los títulos deben ser descriptivos entre 30-60 caracteres'
            })
            
        if long_titles > 0:
            recommendations.append({
                'category': 'Títulos',
                'priority': 'Media',
                'action': f'Acortar {long_titles} títulos que exceden 60 caracteres',
                'details': 'Títulos muy largos se truncan en los resultados de búsqueda'
            })
        
        # Meta description analysis
        no_meta_desc = len([c for c in content_data if not c.get('meta_description')])
        short_meta = len([c for c in content_data if c.get('meta_description') and len(c['meta_description']) < 120])
        long_meta = len([c for c in content_data if c.get('meta_description') and len(c['meta_description']) > 160])
        
        if no_meta_desc > 0:
            recommendations.append({
                'category': 'Meta Descriptions',
                'priority': 'Alta',
                'action': f'Agregar meta descriptions a {no_meta_desc} páginas',
                'details': 'Las meta descriptions mejoran el CTR en resultados de búsqueda'
            })
            
        # Content analysis
        thin_content = len([c for c in content_data if c.get('word_count', 0) < 300])
        if thin_content > 0:
            recommendations.append({
                'category': 'Contenido',
                'priority': 'Alta',
                'action': f'Expandir contenido de {thin_content} páginas con menos de 300 palabras',
                'details': 'Contenido más extenso y valioso mejora el posicionamiento'
            })
        
        # Keyword opportunities
        top_keywords = keyword_analysis.get('top_keywords', [])[:10]
        if top_keywords:
            keyword_list = [f"{kw[0]} ({kw[1]})" for kw in top_keywords]
            recommendations.append({
                'category': 'Palabras Clave',
                'priority': 'Media',
                'action': 'Optimizar contenido para palabras clave principales identificadas',
                'details': f"Keywords principales: {', '.join(keyword_list[:5])}"
            })
        
        # Internal linking
        avg_links = sum([len(c.get('links', [])) for c in content_data]) / pages_analyzed if pages_analyzed > 0 else 0
        if avg_links < 3:
            recommendations.append({
                'category': 'Enlaces Internos',
                'priority': 'Media',
                'action': 'Mejorar estructura de enlaces internos',
                'details': f'Promedio actual: {avg_links:.1f} enlaces por página. Objetivo: 3-10 enlaces relevantes'
            })
        
        return recommendations

    def generate_report(self, content_data, keyword_analysis, recommendations, output_file):
        """Generate markdown SEO report"""
        report = f"""# Plan de Acción SEO - {datetime.now().strftime('%Y-%m-%d')}

## Resumen Ejecutivo

### Páginas Analizadas
- **Total de URLs procesadas**: {len(content_data)}
- **Páginas analizadas exitosamente**: {len([c for c in content_data if c.get('status') != 'failed'])}
- **Errores de acceso**: {len([c for c in content_data if c.get('status') == 'failed'])}

### Palabras Clave Principales
"""
        
        top_keywords = keyword_analysis.get('top_keywords', [])[:15]
        for i, (keyword, count) in enumerate(top_keywords, 1):
            report += f"{i}. **{keyword}** ({count} menciones)\n"
        
        report += "\n### Frases Relevantes\n"
        top_phrases = keyword_analysis.get('top_phrases', [])[:10]
        for i, (phrase, count) in enumerate(top_phrases, 1):
            report += f"{i}. \"{phrase}\" ({count} menciones)\n"
        
        report += "\n## Recomendaciones de Acción\n\n"
        
        # Group recommendations by priority
        high_priority = [r for r in recommendations if r['priority'] == 'Alta']
        medium_priority = [r for r in recommendations if r['priority'] == 'Media']
        low_priority = [r for r in recommendations if r['priority'] == 'Baja']
        
        if high_priority:
            report += "### 🔴 Prioridad Alta (Implementar Inmediatamente)\n\n"
            for i, rec in enumerate(high_priority, 1):
                report += f"**{i}. {rec['category']}: {rec['action']}**\n"
                report += f"- *Detalles*: {rec['details']}\n"
                report += f"- *Impacto*: Alto impacto en SEO\n"
                report += f"- *Esfuerzo*: Técnico\n\n"
        
        if medium_priority:
            report += "### 🟡 Prioridad Media (Implementar en 2-4 semanas)\n\n"
            for i, rec in enumerate(medium_priority, 1):
                report += f"**{i}. {rec['category']}: {rec['action']}**\n"
                report += f"- *Detalles*: {rec['details']}\n"
                report += f"- *Impacto*: Medio impacto en SEO\n"
                report += f"- *Esfuerzo*: Moderado\n\n"
        
        if low_priority:
            report += "### 🟢 Prioridad Baja (Implementar en 1-2 meses)\n\n"
            for i, rec in enumerate(low_priority, 1):
                report += f"**{i}. {rec['category']}: {rec['action']}**\n"
                report += f"- *Detalles*: {rec['details']}\n"
                report += f"- *Impacto*: Bajo impacto en SEO\n"
                report += f"- *Esfuerzo*: Mínimo\n\n"
        
        report += """## Estrategia de Link Building

### 1. Enlaces Internos
- Crear una arquitectura de enlaces internos coherente
- Utilizar anchor text descriptivo con palabras clave relevantes
- Enlazar páginas relacionadas temáticamente
- Implementar breadcrumbs para mejorar la navegación

### 2. Contenido de Calidad
- Crear contenido valioso que genere enlaces naturales
- Desarrollar guías completas sobre temas de nicho
- Publicar estudios de caso y análisis técnicos
- Mantener el contenido actualizado y relevante

### 3. Outreach y Relaciones Públicas
- Identificar sitios web relevantes en el sector
- Contactar con bloggers y periodistas del sector
- Participar en foros y comunidades especializadas
- Colaborar con influencers del sector

### 4. Directorios y Listados
- Registrarse en directorios de calidad del sector
- Completar perfiles en plataformas B2B relevantes
- Mantener información consistente (NAP) en todos los listados

## Métricas de Seguimiento

### KPIs Principales
1. **Posiciones de palabras clave objetivo**
2. **Tráfico orgánico mensual**
3. **Número de páginas indexadas**
4. **Tiempo de permanencia y bounce rate**
5. **Número y calidad de enlaces entrantes**
6. **Conversiones desde tráfico orgánico**

### Herramientas Recomendadas
- Google Search Console (imprescindible)
- Google Analytics 4
- SEMrush o Ahrefs (análisis de competencia)
- Screaming Frog (auditorías técnicas)

## Cronograma de Implementación

### Semana 1-2: Optimizaciones Críticas
- Corregir títulos y meta descriptions faltantes
- Solucionar errores técnicos identificados
- Implementar estructura de datos

### Semana 3-4: Mejoras de Contenido  
- Expandir páginas con contenido delgado
- Optimizar contenido existente para palabras clave
- Mejorar estructura de headings

### Mes 2: Estrategia de Enlaces
- Implementar estrategia de enlaces internos
- Inicio de campaña de outreach
- Crear contenido linkable

### Mes 3+: Optimización Continua
- Monitoreo y ajuste de estrategias
- Análisis de resultados y iteración
- Expansión de contenido basada en datos

---
*Reporte generado automáticamente el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*
"""
        
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(report)
            logger.info(f"Report saved to: {output_file}")
        except Exception as e:
            logger.error(f"Error saving report: {e}")

    def auto_adjust_columns(self, worksheet, max_width=80):
        """Safely auto-adjust column widths, handling merged cells"""
        try:
            for column in worksheet.columns:
                max_length = 0
                # Get column letter from first non-merged cell
                column_letter = None

                for cell in column:
                    # Skip merged cells
                    if isinstance(cell, MergedCell):
                        continue

                    # Get column letter from first valid cell
                    if column_letter is None:
                        column_letter = cell.column_letter

                    # Calculate max length
                    try:
                        cell_length = len(str(cell.value)) if cell.value is not None else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass

                # Apply width adjustment if we found a valid column letter
                if column_letter and max_length > 0:
                    adjusted_width = min(max_length + 2, max_width)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            logger.warning(f"Could not auto-adjust columns: {e}")

    def generate_excel_report(self, content_data, keyword_analysis, recommendations, output_file):
        """Generate comprehensive Excel SEO report with multiple sheets"""
        try:
            wb = Workbook()

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            subheader_font = Font(bold=True, color="000000")
            subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            wrap_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 1. RESUMEN EJECUTIVO SHEET
            ws_summary = wb.create_sheet("Resumen Ejecutivo")

            # Title
            ws_summary.merge_cells('A1:F1')
            ws_summary['A1'] = f"ANÁLISIS SEO - {datetime.now().strftime('%Y-%m-%d')}"
            ws_summary['A1'].font = Font(size=16, bold=True)
            ws_summary['A1'].alignment = center_alignment

            # Summary stats
            successful_pages = len([c for c in content_data if c.get('status') != 'failed'])
            failed_pages = len([c for c in content_data if c.get('status') == 'failed'])

            row = 3
            summary_data = [
                ["Métrica", "Valor"],
                ["Total URLs procesadas", len(content_data)],
                ["Páginas analizadas exitosamente", successful_pages],
                ["Errores de acceso", failed_pages],
                ["Porcentaje de éxito", f"{(successful_pages/len(content_data)*100):.1f}%" if content_data else "0%"],
                ["Palabras clave únicas identificadas", len(keyword_analysis.get('top_keywords', []))],
                ["Recomendaciones generadas", len(recommendations)]
            ]

            for row_data in summary_data:
                for col, value in enumerate(row_data, 1):
                    cell = ws_summary.cell(row=row, column=col, value=value)
                    if row == 3:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = border
                row += 1

            # Auto-adjust column widths
            self.auto_adjust_columns(ws_summary, max_width=50)

            # 2. ANÁLISIS DE PÁGINAS SHEET
            ws_pages = wb.create_sheet("Análisis de Páginas")

            # Headers
            page_headers = [
                "URL", "Estado", "Título", "Longitud Título", "Meta Description",
                "Longitud Meta Desc", "Palabras", "Enlaces Internos", "H1", "H2", "H3"
            ]

            for col, header in enumerate(page_headers, 1):
                cell = ws_pages.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Page data
            for row, content in enumerate(content_data, 2):
                if content.get('status') == 'failed':
                    status = "ERROR"
                    title = content.get('error', 'Error desconocido')
                    title_length = 0
                    meta_desc = ""
                    meta_length = 0
                    word_count = 0
                    links_count = 0
                    h1_count = 0
                    h2_count = 0
                    h3_count = 0
                else:
                    status = "OK"
                    title = content.get('title', '')
                    title_length = len(title)
                    meta_desc = content.get('meta_description', '')
                    meta_length = len(meta_desc)
                    word_count = content.get('word_count', 0)
                    links_count = len(content.get('links', []))
                    headings = content.get('headings', {})
                    h1_count = len(headings.get('h1', []))
                    h2_count = len(headings.get('h2', []))
                    h3_count = len(headings.get('h3', []))

                row_data = [
                    content.get('url', ''),
                    status,
                    title,
                    title_length,
                    meta_desc,
                    meta_length,
                    word_count,
                    links_count,
                    h1_count,
                    h2_count,
                    h3_count
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws_pages.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col in [1, 3, 5]:  # URL, Title, Meta Description columns
                        cell.alignment = wrap_alignment
                    else:
                        cell.alignment = center_alignment

                    # Color coding for issues
                    if col == 2 and value == "ERROR":  # Status column
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif col == 4:  # Title length
                        if value == 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif value < 30 or value > 60:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif col == 6:  # Meta description length
                        if value == 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif value < 120 or value > 160:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif col == 7 and value < 300:  # Word count
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            # Auto-adjust column widths
            self.auto_adjust_columns(ws_pages, max_width=80)

            # 3. PALABRAS CLAVE SHEET
            ws_keywords = wb.create_sheet("Palabras Clave")

            # Headers
            kw_headers = ["Posición", "Palabra Clave", "Frecuencia", "Relevancia"]
            for col, header in enumerate(kw_headers, 1):
                cell = ws_keywords.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Keyword data
            top_keywords = keyword_analysis.get('top_keywords', [])[:50]
            for row, (keyword, count) in enumerate(top_keywords, 2):
                relevance = "Alta" if count >= 10 else "Media" if count >= 5 else "Baja"
                row_data = [row-1, keyword, count, relevance]

                for col, value in enumerate(row_data, 1):
                    cell = ws_keywords.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = center_alignment

                    # Color coding by relevance
                    if col == 4:  # Relevance column
                        if value == "Alta":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value == "Media":
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Auto-adjust column widths
            self.auto_adjust_columns(ws_keywords, max_width=50)

            # 4. FRASES RELEVANTES SHEET
            ws_phrases = wb.create_sheet("Frases Relevantes")

            # Headers
            phrase_headers = ["Posición", "Frase", "Frecuencia", "Longitud"]
            for col, header in enumerate(phrase_headers, 1):
                cell = ws_phrases.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Phrase data
            top_phrases = keyword_analysis.get('top_phrases', [])[:30]
            for row, (phrase, count) in enumerate(top_phrases, 2):
                row_data = [row-1, phrase, count, len(phrase.split())]

                for col, value in enumerate(row_data, 1):
                    cell = ws_phrases.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col == 2:  # Phrase column
                        cell.alignment = wrap_alignment
                    else:
                        cell.alignment = center_alignment

            # Auto-adjust column widths
            self.auto_adjust_columns(ws_phrases, max_width=60)

            # 5. RECOMENDACIONES SHEET
            ws_recommendations = wb.create_sheet("Recomendaciones")

            # Headers
            rec_headers = ["Prioridad", "Categoría", "Acción Requerida", "Detalles", "Impacto Estimado"]
            for col, header in enumerate(rec_headers, 1):
                cell = ws_recommendations.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Recommendations data
            for row, rec in enumerate(recommendations, 2):
                impact = "Alto" if rec['priority'] == 'Alta' else "Medio" if rec['priority'] == 'Media' else "Bajo"
                row_data = [
                    rec['priority'],
                    rec['category'],
                    rec['action'],
                    rec['details'],
                    impact
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws_recommendations.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col in [3, 4]:  # Action and Details columns
                        cell.alignment = wrap_alignment
                    else:
                        cell.alignment = center_alignment

                    # Color coding by priority
                    if col == 1:  # Priority column
                        if value == "Alta":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif value == "Media":
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            # Auto-adjust column widths
            self.auto_adjust_columns(ws_recommendations, max_width=80)

            # 6. PROBLEMAS TÉCNICOS SHEET
            ws_issues = wb.create_sheet("Problemas Técnicos")

            # Headers
            issue_headers = ["URL", "Problema", "Tipo", "Severidad", "Solución Recomendada"]
            for col, header in enumerate(issue_headers, 1):
                cell = ws_issues.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Technical issues data
            row = 2
            for content in content_data:
                url = content.get('url', '')

                # Check for various technical issues
                issues = []

                if content.get('status') == 'failed':
                    issues.append({
                        'problema': 'Error de acceso',
                        'tipo': 'Técnico',
                        'severidad': 'Alta',
                        'solucion': 'Verificar URL y accesibilidad del servidor'
                    })
                else:
                    # Title issues
                    title = content.get('title', '')
                    if not title:
                        issues.append({
                            'problema': 'Título faltante',
                            'tipo': 'SEO',
                            'severidad': 'Alta',
                            'solucion': 'Agregar título único y descriptivo'
                        })
                    elif len(title) < 30:
                        issues.append({
                            'problema': 'Título muy corto',
                            'tipo': 'SEO',
                            'severidad': 'Media',
                            'solucion': 'Expandir título a 30-60 caracteres'
                        })
                    elif len(title) > 60:
                        issues.append({
                            'problema': 'Título muy largo',
                            'tipo': 'SEO',
                            'severidad': 'Media',
                            'solucion': 'Reducir título a máximo 60 caracteres'
                        })

                    # Meta description issues
                    meta_desc = content.get('meta_description', '')
                    if not meta_desc:
                        issues.append({
                            'problema': 'Meta description faltante',
                            'tipo': 'SEO',
                            'severidad': 'Alta',
                            'solucion': 'Agregar meta description de 120-160 caracteres'
                        })
                    elif len(meta_desc) < 120:
                        issues.append({
                            'problema': 'Meta description muy corta',
                            'tipo': 'SEO',
                            'severidad': 'Media',
                            'solucion': 'Expandir meta description a 120-160 caracteres'
                        })
                    elif len(meta_desc) > 160:
                        issues.append({
                            'problema': 'Meta description muy larga',
                            'tipo': 'SEO',
                            'severidad': 'Media',
                            'solucion': 'Reducir meta description a máximo 160 caracteres'
                        })

                    # Content issues
                    word_count = content.get('word_count', 0)
                    if word_count < 300:
                        issues.append({
                            'problema': 'Contenido escaso',
                            'tipo': 'Contenido',
                            'severidad': 'Media',
                            'solucion': 'Expandir contenido a mínimo 300 palabras'
                        })

                    # Heading issues
                    headings = content.get('headings', {})
                    if not headings.get('h1'):
                        issues.append({
                            'problema': 'H1 faltante',
                            'tipo': 'SEO',
                            'severidad': 'Alta',
                            'solucion': 'Agregar etiqueta H1 única por página'
                        })

                # Add issues to sheet
                for issue in issues:
                    row_data = [url, issue['problema'], issue['tipo'], issue['severidad'], issue['solucion']]

                    for col, value in enumerate(row_data, 1):
                        cell = ws_issues.cell(row=row, column=col, value=value)
                        cell.border = border
                        if col in [1, 5]:  # URL and Solution columns
                            cell.alignment = wrap_alignment
                        else:
                            cell.alignment = center_alignment

                        # Color coding by severity
                        if col == 4:  # Severity column
                            if value == "Alta":
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            elif value == "Media":
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                    row += 1

            # Auto-adjust column widths
            self.auto_adjust_columns(ws_issues, max_width=80)

            # Save the workbook
            wb.save(output_file)
            logger.info(f"Excel report saved to: {output_file}")

        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")

    def run_analysis(self, sitemap_path=None, url_list_path=None, output_file="plan-de-accion-seo.md", excel_output=None):
        """Run complete SEO analysis"""
        urls = []

        # Collect URLs from sources
        if sitemap_path:
            urls.extend(self.parse_sitemap(sitemap_path))

        if url_list_path:
            urls.extend(self.load_url_list(url_list_path))

        if not urls:
            logger.error("No URLs found to analyze")
            return

        # Remove duplicates
        urls = list(set(urls))
        logger.info(f"Starting analysis of {len(urls)} unique URLs")

        # Setup domain-based cache
        if urls:
            self._setup_domain_cache(urls[0])

        # Clean expired cache (older than 24 hours) and show statistics
        self._clear_expired_cache()
        cache_stats = self._get_cache_stats()
        logger.info(f"📊 Cache stats: {cache_stats['valid_domains']} valid domains, {cache_stats['old_domains']} old domains, {cache_stats['total_domains']} total")

        # Check if domain needs re-scraping (>= 60 minutes)
        needs_scraping = self._needs_rescraping()

        if not needs_scraping:
            # Cache reciente (< 60 min), mostrar info pero NO usar
            metadata_file = self._get_domain_cache_metadata_file()
            try:
                import json
                with open(metadata_file, 'r') as f:
                    metadata = json.load(f)
                    age_minutes = (datetime.now() - datetime.fromisoformat(metadata['last_scrape'])).total_seconds() / 60

                logger.info(f"")
                logger.info(f"{'='*60}")
                logger.info(f"⚡ USING CACHED DATA - {self.current_domain}")
                logger.info(f"{'='*60}")
                logger.info(f"📦 Last scraped: {age_minutes:.1f} minutes ago")
                logger.info(f"📄 Cached URLs: {metadata.get('url_count', 0)}")
                logger.info(f"⏱️  Cache valid for: {self.cache_rescrape_minutes - age_minutes:.0f} more minutes")
                logger.info(f"{'='*60}")
                logger.info(f"")
                logger.info(f"⚡ Loading data from cache to generate Excel report...")

                # Load all cached data instead of scraping
                for i, url in enumerate(urls, 1):
                    logger.info(f"📦 Loading from cache {i}/{len(urls)}: {url}")
                    content = self.extract_content(url)  # Will use cache automatically
                    if content:
                        self.content_data.append(content)

                logger.info(f"✅ Loaded {len(self.content_data)} pages from cache")
                logger.info(f"")

                # Continue to generate Excel with cached data (don't return here)
                needs_scraping = False
            except:
                needs_scraping = True

        # Si necesitamos scrapear (caché expirado o no existe)
        if needs_scraping:
            logger.info(f"🌐 Starting fresh scraping for {self.current_domain} (required every 60 minutes)")

            # Extract content from URLs (using cache when available)
            cached_count = 0
            scraped_count = 0

            for i, url in enumerate(urls, 1):
                logger.info(f"Processing URL {i}/{len(urls)}: {url}")

                # Check if coming from cache
                is_cached = self._load_from_cache(url) is not None

                content = self.extract_content(url)
                self.content_data.append(content)

                if is_cached:
                    cached_count += 1
                else:
                    scraped_count += 1

                if self.delay > 0 and not is_cached:  # No delay for cached content
                    time.sleep(self.delay)

            # Save domain cache metadata after scraping
            if scraped_count > 0:
                self._save_domain_cache_metadata(len(urls))
        else:
            # Already loaded from cache above
            cached_count = len(self.content_data)
            scraped_count = 0

        # Show cache performance summary
        logger.info(f"")
        logger.info(f"{'='*60}")
        logger.info(f"📈 CACHE PERFORMANCE SUMMARY - {self.current_domain}")
        logger.info(f"{'='*60}")
        logger.info(f"✅ URLs loaded from cache: {cached_count} ({cached_count/len(urls)*100:.1f}%)")
        logger.info(f"🌐 URLs scraped fresh: {scraped_count} ({scraped_count/len(urls)*100:.1f}%)")
        logger.info(f"⚡ Time saved: ~{cached_count * 2:.0f} seconds (approx)")
        logger.info(f"💡 Next scrape required in: {self.cache_rescrape_minutes} minutes")
        logger.info(f"🗑️  Files auto-delete after: {self.cache_cleanup_hours} hours")
        logger.info(f"{'='*60}")
        logger.info(f"")

        # Analyze keywords
        keyword_analysis = self.analyze_keywords(self.content_data)

        # Generate recommendations
        recommendations = self.generate_seo_recommendations(self.content_data, keyword_analysis)

        # Generate markdown report
        self.generate_report(self.content_data, keyword_analysis, recommendations, output_file)

        # Generate Excel report if requested
        if excel_output:
            self.generate_excel_report(self.content_data, keyword_analysis, recommendations, excel_output)
            logger.info(f"Analysis complete! Reports saved to: {output_file} and {excel_output}")
        else:
            logger.info(f"Analysis complete! Report saved to: {output_file}")

def main():
    """Main function with dual operation mode support"""
    parser = argparse.ArgumentParser(description='SEO Content Generator - Advanced SEO Analysis and Content Creation')
    parser.add_argument('--brief', help='Path to business brief file (.txt, .docx, .pdf) for new site mode')
    parser.add_argument('--sitemap', help='Path to sitemap.xml file or URL for existing site mode')
    parser.add_argument('--urls', help='Path to text file with URLs to analyze (existing site mode)')
    parser.add_argument('--output', default='brief-seo-completo.md', help='Output markdown file name')
    parser.add_argument('--excel', default='analisis-seo-completo.xlsx', help='Output Excel file name (.xlsx)')
    parser.add_argument('--delay', type=float, default=1.0, help='Delay between requests (seconds)')
    parser.add_argument('--mode', choices=['new', 'existing'], help='Operation mode: new (brief) or existing (sitemap)')
    parser.add_argument('--clear-cache', action='store_true', help='Clear all cached data before analysis')
    parser.add_argument('--cache-duration', type=int, default=60, help='Cache duration in minutes (default: 60)')
    parser.add_argument('--phase', type=int, choices=[1, 2], help='Execution phase: 1=Generate basic Excel (user fills keywords), 2=Generate AI content using user-defined keywords')

    args = parser.parse_args()

    generator = SEOContentGenerator(delay=args.delay)

    # Set cache duration from arguments
    generator.cache_duration_minutes = args.cache_duration

    # Clear cache if requested
    if args.clear_cache:
        import shutil
        if generator.cache_dir.exists():
            shutil.rmtree(generator.cache_dir)
            generator.cache_dir.mkdir(exist_ok=True)
            logger.info("🗑️  Cache cleared successfully")

    try:
        # Check if using 2-phase workflow
        if args.phase:
            if args.phase == 1:
                # PHASE 1: Generate basic Excel without AI recommendations
                print(f"\n📊 FASE 1: GENERACIÓN EXCEL BÁSICO")
                print("="*60)
                print("📝 Se generará Excel con datos extraídos del sitio web")
                print("⚠️  La columna 'Palabra clave principal' quedará VACÍA para que la completes manualmente")
                print("⚠️  Las columnas de recomendaciones IA quedarán VACÍAS")
                print("")

                if not args.sitemap and not args.urls:
                    print("❌ Error: Para Fase 1 debe proporcionar --sitemap o --urls")
                    return

                # Run analysis and scraping
                generator.analyze_website_from_sitemap(
                    args.sitemap,
                    args.urls,
                    args.output,
                    args.excel
                )

                # Generate Phase 1 Excel (without AI recommendations)
                generator.generate_phase1_excel(args.excel)

                print(f"\n✅ FASE 1 COMPLETADA")
                print("="*60)
                print(f"📊 Excel básico generado: {args.excel}")
                print("")
                print("📝 PRÓXIMOS PASOS:")
                print("   1. Abre el archivo Excel generado")
                print("   2. Edita la columna 7 'Palabra clave principal' en hojas Blog y SEO On-Page")
                print("   3. Define keywords específicas (ej: 'fisioterapia deportiva Madrid')")
                print("   4. Guarda el archivo Excel")
                print("")
                print("🚀 EJECUTA FASE 2:")
                print(f"   python seo_analyzer.py --excel {args.excel} --phase 2")
                print("")
                return

            elif args.phase == 2:
                # PHASE 2: Read Excel with user-defined keywords and generate AI content
                print(f"\n🤖 FASE 2: GENERACIÓN CONTENIDO IA")
                print("="*60)

                if not args.excel:
                    print("❌ Error: Para Fase 2 debe proporcionar --excel con el archivo de Fase 1")
                    return

                if not os.path.exists(args.excel):
                    print(f"❌ Error: Archivo Excel no encontrado: {args.excel}")
                    print(f"   Verifica que el archivo existe y la ruta es correcta")
                    return

                # Generate Phase 2 Excel (AI recommendations based on user keywords)
                generator.generate_phase2_excel(args.excel)

                print(f"\n✅ FASE 2 COMPLETADA")
                print("="*60)
                print(f"📊 Excel con recomendaciones IA: {args.excel}")
                print("")
                return

        # Original workflow (non-phased)
        # Determine operation mode
        if args.mode:
            mode = 'new_site' if args.mode == 'new' else 'existing_site'
        elif args.brief:
            mode = 'new_site'
        elif args.sitemap or args.urls:
            mode = 'existing_site'
        else:
            # Interactive mode selection
            mode = generator.select_operation_mode()

        if mode == 'new_site':
            # New Site Mode - Process brief and generate content
            print(f"\n🚀 MODO: SITIO NUEVO")
            print("="*60)

            # Process brief file
            brief_content = generator.process_brief_file(args.brief)
            if not brief_content:
                return

            # Validate and complete brief
            generator.validate_brief_completeness()

            # Extract keywords for competitive research
            activity = generator.business_data.get('general_info', {}).get('actividad_principal', '')
            location = generator.business_data.get('general_info', {}).get('cobertura_geografica', '')

            research_keywords = [activity]
            if location:
                research_keywords.append(f"{activity} {location}")

            # Perform competitive research
            if research_keywords[0]:  # Only if we have business activity
                generator.search_competitors(research_keywords)
                generator.analyze_serp_patterns()

            # Generate SEO content
            generator.generate_seo_content()

            # Generate outputs
            generator.generate_complete_brief_markdown(args.output)
            generator.generate_expanded_excel_report(args.excel)

            print(f"\n🎉 PROCESO COMPLETADO")
            print(f"📝 Brief completo: {args.output}")
            print(f"📊 Análisis Excel: {args.excel}")

        else:
            # Existing Site Mode - Enhanced analysis
            print(f"\n🔍 MODO: SITIO EXISTENTE")
            print("="*60)

            if not args.sitemap and not args.urls:
                print("❌ Error: Para sitio existente debe proporcionar --sitemap o --urls")
                return

            # Run enhanced analysis
            generator.run_enhanced_analysis(
                sitemap_path=args.sitemap,
                url_list_path=args.urls,
                output_file=args.output,
                excel_output=args.excel
            )

    except KeyboardInterrupt:
        print("\n\n👋 Proceso cancelado por el usuario.")
    except Exception as e:
        logger.error(f"Error during execution: {e}")
        print(f"\n❌ Error durante la ejecución: {e}")
    finally:
        # Close WebDriver if initialized
        if generator.driver:
            try:
                generator.driver.quit()
                logger.info("WebDriver closed successfully")
            except:
                pass

if __name__ == "__main__":
    main()